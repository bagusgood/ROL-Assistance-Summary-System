import requests
import pandas as pd
import plotly.express as px
import plotly
import json
from flask import send_file
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import re
from plotly.utils import PlotlyJSONEncoder
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break
from openpyxl.styles import Alignment, Border, Side
from openpyxl.styles import PatternFill
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.lib import colors
from reportlab.pdfbase.ttfonts import TTFont
import locale
from flask import Flask, render_template_string, request, redirect, url_for, session, flash
import logging, os, tempfile
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import plotly.graph_objects as go
import folium
from folium.plugins import MarkerCluster
import matplotlib.pyplot as plt
from io import StringIO
from flask import Flask, render_template, jsonify
import os
from flask import Flask, render_template, request, jsonify, send_file, send_from_directory, redirect, url_for
import pandas as pd
import numpy as np
from geopy.distance import geodesic
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = "rahasia_super"  # ganti dengan secret key lebih kuat

# ======= USER LOGIN =======
USERS = {
    "balmon_mataram": "rahasia_umum",   # username: password
    "username": "password"
}

# ================== CONFIG INVOICE ==================
INVOICE_BASE_URL = "https://dendaadministratif.postel.go.id"
INVOICE_LOGIN_URL = f"{INVOICE_BASE_URL}/auth/login"
INVOICE_DATA_URL = f"{INVOICE_BASE_URL}/application/invoice/get/data/management"

INVOICE_USERNAME = "pic1_upt_mataram"
INVOICE_PASSWORD = "password"

UPLOAD_FOLDER = "uploads"
STATIC_FOLDER = "static"

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(STATIC_FOLDER, exist_ok=True)

import matplotlib
matplotlib.use("Agg")

import os
from flask import Flask, request, render_template_string, redirect, url_for


# Konfigurasi upload
UPLOAD_FOLDER = 'data'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
EXCEL_FILE = 'data/hasil_klasifikasi_gabungan.xlsx'

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Load data Excel
def load_bts_data():
    try:
        # Cek apakah file Excel ada
        if not os.path.exists(EXCEL_FILE):
            print(f"File {EXCEL_FILE} tidak ditemukan")
            return pd.DataFrame()
        
        # Baca file Excel
        df = pd.read_excel(EXCEL_FILE)
        
        # Bersihkan nama kolom
        df.columns = df.columns.str.strip()
        
        # Pilih kolom yang relevan
        required_columns = [
            'CLNT_ID', 'CLNT_NAME', 'STN_NAME', 'LAT_VAL', 'LONG_VAL',
            'LAT_DEG', 'LAT_MIN', 'LAT_SEC', 'LAT_DIR_IND',
            'LONG_DEG', 'LONG_MIN', 'LONG_SEC', 'LONG_DIR_IND',
            'cluster_id', 'SERVICE', 'FREQ', 'CITY', 'DISTRICT', 'DIST_RROR (m)'
        ]
        
        if 'LAT_VAL' not in df.columns:
            df['LAT_VAL'] = None
            df.to_excel(EXCEL_FILE)
            
        if 'LONG_VAL' not in df.columns:
            df['LONG_VAL'] = None
            df.to_excel(EXCEL_FILE)
            
        if 'DIST_RROR (m)' not in df.columns:
            df['DIST_RROR (m)'] = None
            df.to_excel(EXCEL_FILE)
        
        if 'cluster_id' not in df.columns:
            df['cluster_id'] = None
            df.to_excel(EXCEL_FILE)
        
        # Filter hanya kolom yang ada di dataframe
        available_columns = [col for col in required_columns if col in df.columns]
        df = df[available_columns]
        
        
        # Konversi koordinat DMS ke decimal degrees untuk koordinat resmi
        def dms_to_decimal(degrees, minutes, seconds, direction):
            try:
                decimal = float(degrees) + float(minutes)/60 + float(seconds)/3600
                if direction in ['S', 'W']:
                    decimal = -decimal
                return decimal
            except:
                return None
        
        # Buat kolom untuk koordinat resmi
        df['OFFICIAL_LAT'] = df.apply(lambda row: dms_to_decimal(
            row.get('LAT_DEG', 0), 
            row.get('LAT_MIN', 0), 
            row.get('LAT_SEC', 0), 
            row.get('LAT_DIR_IND', 'N')
        ) if pd.notnull(row.get('LAT_DEG')) else None, axis=1)
        
        df['OFFICIAL_LONG'] = df.apply(lambda row: dms_to_decimal(
            row.get('LONG_DEG', 0), 
            row.get('LONG_MIN', 0), 
            row.get('LONG_SEC', 0), 
            row.get('LONG_DIR_IND', 'E')
        ) if pd.notnull(row.get('LONG_DEG')) else None, axis=1)
        
        # Isi data yang kosong - gunakan koordinat resmi jika LAT_VAL/LONG_VAL kosong
        df['LAT_VAL'] = pd.to_numeric(df['LAT_VAL'], errors='coerce')
        df['LONG_VAL'] = pd.to_numeric(df['LONG_VAL'], errors='coerce')
        
        # Untuk marker di peta: SELALU gunakan koordinat resmi (DMS)
        df['MARKER_LAT'] = df['OFFICIAL_LAT']  
        df['MARKER_LNG'] = df['OFFICIAL_LONG']  
        
        # Kolom display untuk UI (sama dengan marker)
        df['DISPLAY_LAT'] = df['MARKER_LAT']
        df['DISPLAY_LONG'] = df['MARKER_LNG']
        
        # Tentukan status berdasarkan DIST_RROR (m)
        def determine_status(row):
            if 'DIST_RROR (m)' in row.index:
                dist_error = row['DIST_RROR (m)']
                if pd.notnull(dist_error):
                    try:
                        # Coba konversi ke float
                        error_str = str(dist_error).strip()
                        if error_str in ['', 'nan', 'none', 'null']:
                            return "belum"  # Belum dihitung
                        
                        try:
                            error_value = float(error_str)
                            if error_value <= 14:
                                return "benar"
                            else:
                                return "salah"
                        except ValueError:
                            import re
                            numbers = re.findall(r'\d+\.?\d*', error_str)
                            if numbers:
                                error_value = float(numbers[0])
                                if error_value <= 14:
                                    return "benar"
                                else:
                                    return "salah"
                            else:
                                return "salah"
                    except:
                        return "belum"
            return "belum"
        
        df['DATA_STATUS'] = df.apply(determine_status, axis=1)
        
        # Hitung error distance jika ada koordinat aktual dan resmi
        def calculate_error(row):
            try:
                if pd.notnull(row['LAT_VAL']) and pd.notnull(row['LONG_VAL']) and \
                   pd.notnull(row['OFFICIAL_LAT']) and pd.notnull(row['OFFICIAL_LONG']):
                    coord1 = (row['LAT_VAL'], row['LONG_VAL'])
                    coord2 = (row['OFFICIAL_LAT'], row['OFFICIAL_LONG'])
                    return geodesic(coord1, coord2).meters
                return None
            except:
                return None
        
        df['CALCULATED_ERROR_M'] = df.apply(calculate_error, axis=1)
        
        # Format untuk JSON
        df['id'] = df.index
        df['name'] = df['STN_NAME'].fillna('Unknown Tower')
        df['lat'] = df['DISPLAY_LAT']  
        df['lng'] = df['DISPLAY_LONG']  
        df['actual_lat'] = df['LAT_VAL']  
        df['actual_lng'] = df['LONG_VAL'] 
        df['official_lat'] = df['OFFICIAL_LAT']
        df['official_lng'] = df['OFFICIAL_LONG']
        df['cluster'] = df['cluster_id'].fillna(0).astype(int)
        df['provider'] = df['CLNT_NAME'].fillna('Unknown')
        
        # Statistik
        total_towers = len(df)
        benar_diambil = len(df[df['DATA_STATUS'] == 'benar'])
        salah_diambil = len(df[df['DATA_STATUS'] == 'salah'])
        belum_diambil = len(df[df['DATA_STATUS'] == 'belum'])
        
        return df
    except Exception as e:
        print(f"Error loading data: {e}")
        return pd.DataFrame()
    
# Fungsi untuk update data di Excel
def update_tower_data(tower_id, measured_error, measured_lat=None, measured_lng=None):
    try:
        # Baca semua sheet
        excel_data = pd.read_excel(EXCEL_FILE, sheet_name=None)
        
        # Ambil sheet utama
        main_sheet_name = 'Sheet1'
        if main_sheet_name in excel_data:
            df = excel_data[main_sheet_name]
            
            # Pastikan tower_id valid
            if tower_id >= 0 and tower_id < len(df):
                # Update DIST_RROR (m)
                if 'DIST_RROR (m)' in df.columns:
                    df.at[tower_id, 'DIST_RROR (m)'] = measured_error
                
                # Update LAT_VAL dan LONG_VAL jika diberikan
                if measured_lat is not None and 'LAT_VAL' in df.columns:
                    df.at[tower_id, 'LAT_VAL'] = measured_lat
                
                if measured_lng is not None and 'LONG_VAL' in df.columns:
                    df.at[tower_id, 'LONG_VAL'] = measured_lng
                
                # Simpan kembali ke Excel
                with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
                    for sheet_name, sheet_df in excel_data.items():
                        if sheet_name == main_sheet_name:
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
                        else:
                            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                print(f"Updated tower {tower_id}: error={measured_error}m")
                return True
            else:
                print(f"Invalid tower_id: {tower_id}")
                return False
        else:
            print(f"Sheet {main_sheet_name} not found")
            return False
    except Exception as e:
        print(f"Error updating Excel: {e}")
        return False

bts_data = load_bts_data()


@app.route("/run_bts", methods=["GET", "POST"])
def run_bts():
    """Halaman utama dengan peta dan semua tower"""
    if not os.path.exists(EXCEL_FILE):
        # Redirect ke halaman upload jika file tidak ditemukan
        return redirect(url_for('upload_file'))
    try:
         # Ambil informasi file
        file_info = {
            'exists': os.path.exists(EXCEL_FILE),
            'size': os.path.getsize(EXCEL_FILE) if os.path.exists(EXCEL_FILE) else 0
        }
       
        
        # Ambil data unik untuk dropdown
        towers = bts_data[['id', 'name', 'cluster', 'provider', 'DATA_STATUS']].to_dict('records')
        clusters = sorted(bts_data['cluster'].unique().tolist())
        providers = sorted(bts_data['provider'].unique().tolist())
        
        # Hitung statistik
        total_towers = len(bts_data)
        benar_diambil = len(bts_data[bts_data['DATA_STATUS'] == 'benar'])
        salah_diambil = len(bts_data[bts_data['DATA_STATUS'] == 'salah'])
        belum_diambil = len(bts_data[bts_data['DATA_STATUS'] == 'belum'])
        
        
        # Konversi data ke GeoJSON format
        features = []
        for idx, row in bts_data.iterrows():
            # Gunakan display coordinates (jika kosong, gunakan official)
            display_lat = row['lat']
            display_lng = row['lng']
            
            # Jika masih kosong, skip (tapi seharusnya sudah diisi di load_bts_data)
            if pd.isnull(display_lat) or pd.isnull(display_lng):
                continue
            
            # Ambil nilai dist_error
            dist_error_value = None
            if 'dist_error' in row and pd.notnull(row['dist_error']):
                try:
                    dist_error_value = float(row['dist_error'])
                except:
                    dist_error_value = None
            
            feature = {
                "type": "Feature",
                "properties": {
                    "id": int(row['id']),
                    "name": str(row['name']),
                    "cluster": int(row['cluster']),
                    "provider": str(row['provider']),
                    "status": str(row['DATA_STATUS']),
                    "error_m": float(row['CALCULATED_ERROR_M']) if pd.notnull(row['CALCULATED_ERROR_M']) else None,
                    "dist_error": dist_error_value,
                    "has_coords": pd.notnull(row['actual_lat']) and pd.notnull(row['actual_lng'])
                },
                "geometry": {
                    "type": "Point",
                    "coordinates": [float(display_lng), float(display_lat)]
                }
            }
            features.append(feature)
        
        geojson_data = {
            "type": "FeatureCollection",
            "features": features
        }
        
        return render_template('index.html', 
                             towers=towers,
                             clusters=clusters,
                             providers=providers,
                             total_towers=total_towers,
                             benar_diambil=benar_diambil,
                             salah_diambil=salah_diambil,
                             belum_diambil=belum_diambil,
                             file_info = file_info,
                             geojson_data=json.dumps(geojson_data))
    except Exception as e:
        print(f"Error in index route: {e}")
        return f"Error: {str(e)}", 500
    
# Route untuk upload file
@app.route('/upload', methods=['GET', 'POST'])
def upload_file():
     # Ambil informasi file
    file_info = {
        'exists': os.path.exists(EXCEL_FILE),
        'size': os.path.getsize(EXCEL_FILE) if os.path.exists(EXCEL_FILE) else 0
    }
    """Halaman untuk upload file Excel"""
    if request.method == 'POST':
        # Cek apakah ada file yang diupload
        if 'excel_file' not in request.files:
            return jsonify({"error": "Tidak ada file yang dipilih"}), 400
        
        file = request.files['excel_file']
        
        # Jika user tidak memilih file
        if file.filename == '':
            return jsonify({"error": "Tidak ada file yang dipilih"}), 400
        
        if file and allowed_file(file.filename):
            try:
                # Simpan file
                filename = secure_filename('hasil_klasifikasi_gabungan.xlsx')
                filepath = os.path.join("data/" + filename)
                file.save(filepath)
                
                # Reload data
                global bts_data
                bts_data = load_bts_data()
                
                # Hitung statistik baru
                total_towers = len(bts_data)
                benar_diambil = len(bts_data[bts_data['DATA_STATUS'] == 'benar'])
                salah_diambil = len(bts_data[bts_data['DATA_STATUS'] == 'salah'])
                belum_diambil = len(bts_data[bts_data['DATA_STATUS'] == 'belum'])
                
                return jsonify({
                    "success": True,
                    "message": "File berhasil diupload dan data telah direload",
                    "filename": filename,
                    "total_towers": total_towers,
                    "benar_diambil": benar_diambil,
                    "salah_diambil": salah_diambil,
                    "belum_diambil": belum_diambil
                })
                
            except Exception as e:
                return jsonify({"error": f"Error saat menyimpan file: {str(e)}"}), 500
        
        return jsonify({"error": "Format file tidak diizinkan. Hanya file Excel (.xlsx, .xls)"}), 400
    
    # GET request - tampilkan halaman upload
    return render_template('upload.html',
                           file_info=file_info)
    
# Route untuk mendownload file Excel yang ada di folder data
@app.route('/download-excel')
def download_excel_bts():
    """Download file Excel saat ini dari folder data"""
    try:
        # Cek apakah file ada
        if not os.path.exists(EXCEL_FILE):
            return jsonify({"error": "File Excel tidak ditemukan. Silakan upload file terlebih dahulu."}), 404
        
        # Tentukan nama file untuk download
        download_name = 'hasil_analisis_bts_tower.xlsx'
        
        # Kirim file sebagai attachment
        return send_file(
            EXCEL_FILE,
            as_attachment=True,
            download_name=download_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        print(f"Error downloading Excel: {e}")
        return jsonify({"error": str(e)}), 500
    
@app.route('/download-excel-<string:filename>')
def download_excel_custom(filename):
    """Download file Excel dengan nama custom"""
    try:
        # Cek apakah file ada
        if not os.path.exists(EXCEL_FILE):
            return jsonify({"error": "File Excel tidak ditemukan"}), 404
        
        # Validasi nama file untuk keamanan
        if not filename.endswith('.xlsx'):
            filename = filename + '.xlsx'
        
        # Kirim file sebagai attachment
        return send_file(
            EXCEL_FILE,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        print(f"Error downloading Excel with custom name: {e}")
        return jsonify({"error": str(e)}), 500
    
# Route untuk informasi file saat ini
@app.route('/api/current-file-info')
def get_current_file_info():
    """API untuk mendapatkan informasi file Excel saat ini"""
    try:
        if os.path.exists(EXCEL_FILE):
            file_stat = os.stat(EXCEL_FILE)
            return jsonify({
                'exists': True,
                'filename': 'hasil_klasifikasi_gabungan.xlsx',
                'size': file_stat.st_size,
                'modified': file_stat.st_mtime,
                'last_modified': file_stat.st_mtime,
                'path': EXCEL_FILE
            })
        else:
            return jsonify({'exists': False})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    


@app.route('/api/towers')
def get_towers():
    """API untuk mendapatkan semua tower"""
    towers = bts_data[['id', 'name', 'lat', 'lng', 'cluster', 'provider', 'DATA_STATUS']].to_dict('records')
    return jsonify(towers)

@app.route('/api/tower/<int:tower_id>')
def get_tower(tower_id):
    """API untuk mendapatkan data tower tertentu"""
    if tower_id >= 0 and tower_id < len(bts_data):
        tower_data = bts_data.iloc[tower_id].to_dict()
        return jsonify(tower_data)
    return jsonify({"error": "Tower not found"}), 404

@app.route('/api/update-tower', methods=['POST'])
def update_tower():
    """API untuk update data tower setelah pengukuran"""
    try:
        data = request.json
        tower_id = data.get('tower_id')
        measured_error = data.get('measured_error')
        measured_lat = data.get('measured_lat')
        measured_lng = data.get('measured_lng')
        
        if tower_id is None or measured_error is None:
            return jsonify({"error": "Missing required parameters"}), 400
        
        # Update di Excel file
        success = update_tower_data(
            tower_id=tower_id,
            measured_error=measured_error,
            measured_lat=measured_lat,
            measured_lng=measured_lng
        )
        
        if success:
            # Reload data
            global bts_data
            bts_data = load_bts_data()
            
            return jsonify({
                "success": True,
                "message": "Data berhasil disimpan",
                "tower_id": tower_id,
                "new_error": measured_error
            })
        else:
            return jsonify({"error": "Gagal menyimpan data"}), 500
            
    except Exception as e:
        print(f"Error updating tower: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/search')
def search_towers():
    """API untuk search tower"""
    query = request.args.get('q', '').lower()
    cluster = request.args.get('cluster', '')
    provider = request.args.get('provider', '')
    
    filtered = bts_data.copy()
    
    if query:
        filtered = filtered[filtered['name'].str.lower().str.contains(query, na=False) |
                           filtered['provider'].str.lower().str.contains(query, na=False)]
    
    if cluster:
        filtered = filtered[filtered['cluster'] == int(cluster)]
    
    if provider:
        filtered = filtered[filtered['provider'] == provider]
    
    results = filtered[['id', 'name', 'lat', 'lng', 'cluster', 'provider']].to_dict('records')
    return jsonify(results)

@app.route('/calculate-error', methods=['POST'])
def calculate_error():
    """API untuk menghitung error distance"""
    data = request.json
    tower_id = data.get('tower_id')
    clicked_lat = data.get('lat')
    clicked_lng = data.get('lng')
    
    if not all([tower_id, clicked_lat, clicked_lng]):
        return jsonify({"error": "Missing parameters"}), 400
    
    tower = bts_data[bts_data['id'] == tower_id]
    if tower.empty:
        return jsonify({"error": "Tower not found"}), 404
    
    tower = tower.iloc[0]
    
    # Koordinat resmi dari data
    official_lat = tower['OFFICIAL_LAT']
    official_lng = tower['OFFICIAL_LONG']
    
    if pd.isnull(official_lat) or pd.isnull(official_lng):
        return jsonify({"error": "Official coordinates not available"}), 400
    
    # Hitung distance antara koordinat resmi dan yang diklik
    coord1 = (official_lat, official_lng)
    coord2 = (clicked_lat, clicked_lng)
    
    try:
        distance_m = geodesic(coord1, coord2).meters
        
        # Hitung juga error dari data aktual
        actual_error = None
        if pd.notnull(tower['CALCULATED_ERROR_M']):
            actual_error = tower['CALCULATED_ERROR_M']
        
        return jsonify({
            "success": True,
            "tower_name": str(tower['name']),
            "official_coords": {
                "lat": float(official_lat),
                "lng": float(official_lng)
            },
            "clicked_coords": {
                "lat": float(clicked_lat),
                "lng": float(clicked_lng)
            },
            "error_distance_m": round(distance_m, 2),
            "actual_error_m": round(actual_error, 2) if actual_error else None,
            "cluster": int(tower['cluster']),
            "provider": str(tower['provider'])
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/tower/<int:tower_id>')
def tower_detail(tower_id):
    """Halaman detail tower"""
    tower = bts_data[bts_data['id'] == tower_id]
    if tower.empty:
        return "Tower not found", 404
    
    tower_data = tower.iloc[0].to_dict()
    
    # Konversi data ke format yang aman untuk JavaScript
    for key, value in tower_data.items():
        if isinstance(value, (np.integer, np.floating)):
            tower_data[key] = float(value)
        elif pd.isnull(value):
            tower_data[key] = None
    
    return render_template('result.html', tower=tower_data)

@app.route('/api/stats')
def get_stats():
    """API untuk mendapatkan statistik data"""
    total = len(bts_data)
    benar = len(bts_data[bts_data['DATA_STATUS'] == 'benar'])
    salah = len(bts_data[bts_data['DATA_STATUS'] == 'salah'])
    belum = len(bts_data[bts_data['DATA_STATUS'] == 'belum'])
    
    return jsonify({
        'total': total,
        'benar': benar,
        'salah': salah,
        'belum': belum,
        'persen_benar': round(benar/total*100, 1) if total > 0 else 0,
        'persen_salah': round(salah/total*100, 1) if total > 0 else 0,
        'persen_belum': round(belum/total*100, 1) if total > 0 else 0
    })

# ======================================================
# CONFIG
# ======================================================
UPLOAD_FOLDER = "uploads"
STATIC_FOLDER = "static"

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(STATIC_FOLDER, exist_ok=True)


CURRENT_SPECTRUM = None
CURRENT_REKAP = None
CURRENT_FILE_TYPE = None
CURRENT_DATASET_NAME = None

# ======================================================
# FREQUENCY RANGES
# ======================================================
FREQ_RANGES = {
    "87–108 MHz": (87e6, 108e6),
    "108–137 MHz": (108e6, 137e6),
    "137–174 MHz": (137e6, 174e6),
    "174–230 MHz": (174e6, 230e6),
    "300–430 MHz": (300e6, 430e6),
    "430–460 MHz": (430e6, 460e6),
    "460–470 MHz": (460e6, 470e6),
    "478–806 MHz": (478e6, 806e6),
    "806–880 MHz": (806e6, 880e6),
    "925–960 MHz": (925e6, 960e6),
    "1427–1518 MHz": (1427e6, 1518e6),
    "1805–1880 MHz": (1805e6, 1880e6),
    "2110–2170 MHz": (2110e6, 2170e6),
    "2170–2200 MHz": (2170e6, 2200e6),
    "2300–2400 MHz": (2300e6, 2400e6)
}

# ======================================================
# LOAD & CLEAN REKAP (FROM UPLOAD)
# ======================================================
def load_rekap(path):
    df = pd.read_csv(path)
    df.columns = df.columns.str.strip()

    # ===== BUANG SEMUA LEVEL LAMA =====
    for c in df.columns:
        if "level" in c.lower():
            df = df.drop(columns=c)

    # ===== NORMALISASI FREKUENSI =====
    df["Frekuensi"] = (
        df["Frekuensi"]
        .astype(str)
        .str.replace(",", ".", regex=False)
        .str.replace(r"[^0-9\.]", "", regex=True)
    )
    
    df["Frekuensi"] = pd.to_numeric(df["Frekuensi"], errors="coerce")
    df = df.dropna(subset=["Frekuensi"])
    df = df.drop_duplicates(subset=["Frekuensi"])
    df = df.sort_values("Frekuensi").reset_index(drop=True)

    # ===== TAMBAHKAN LEVEL BARU (KOSONG) =====
    df["Level (dBµV/m)"] = ""

    # ===== REBUILD NO =====
    df["No"] = range(1, len(df) + 1)

    df.to_csv("rekap_baru.csv", index=False)
    return df



def attach_level_from_spectrum(df_rekap, df_spec):
    spec_f = df_spec["Frequency (Hz)"].values
    spec_l = df_spec["Level (dBµV/m)"].values

    levels = []

    for f_mhz in df_rekap["Frekuensi"].astype(float).values:
        f_hz = f_mhz * 1e6
        idx = (abs(spec_f - f_hz)).argmin()
        levels.append(spec_l[idx])

    df_rekap["Level (dBµV/m)"] = levels
    return df_rekap



# ======================================================
# LOAD SPECTRUM
# ======================================================
def load_spectrum(filepath, file_type=None):
    import pandas as pd
    import re
    from io import StringIO

    # ==================================================
    # 1. KHUSUS LS TELECOM (RAW EXPORT)
    # ==================================================
    if file_type == "LS TELECOM":
        try:
            df = pd.read_csv(
                filepath,
                sep=";",
                skiprows=11,
                encoding="latin1",
                engine="python"
            )

            # Buang 3 kolom metadata awal
            df = df.iloc[:, 3:]

            if df.empty:
                raise ValueError("Data LS Telecom kosong setelah buang metadata")

            # Transpose: kolom frekuensi → baris
            df_t = df.T.reset_index()
            df_t.rename(columns={"index": "Frequency (Hz)"}, inplace=True)

            # Parsing Frequency
            df_t["Frequency (Hz)"] = (
                df_t["Frequency (Hz)"]
                .astype(str)
                .str.replace(",", ".", regex=False)
                .str.replace(r"[^0-9\.]", "", regex=True)
            )
            df_t["Frequency (Hz)"] = pd.to_numeric(
                df_t["Frequency (Hz)"], errors="coerce"
            )

            # Semua level numerik
            for col in df_t.columns[1:]:
                df_t[col] = pd.to_numeric(df_t[col], errors="coerce")

            # Envelope spectrum (MAX)
            df_t["Level (dBµV/m)"] = df_t.iloc[:, 1:].max(axis=1)

            df_final = df_t[["Frequency (Hz)", "Level (dBµV/m)"]].dropna()

            if df_final.empty:
                raise ValueError("Data LS Telecom kosong setelah parsing")

            return df_final

        except Exception as e:
            raise ValueError(f"Gagal membaca LS TELECOM: {e}")

    # ==================================================
    # 2. GENERIC SPECTRUM (TCI, ARGUS, DLL)
    # ==================================================
    with open(filepath, "r", encoding="latin1", errors="ignore") as f:
        lines = f.readlines()

    if not lines:
        raise ValueError("File spectrum kosong")

    # -----------------------------
    # Deteksi delimiter awal (sep=)
    # -----------------------------
    sep = ","
    start_row = 0

    if lines[0].lower().startswith("sep="):
        sep = lines[0].strip().split("=", 1)[1]
        start_row = 1

    # -----------------------------
    # Cari header Frequency
    # -----------------------------
    header_row = None
    for i in range(start_row, len(lines)):
        line = lines[i].lower()

        if "freq" in line and (
            "level" in line or "field" in line or "strength" in line
        ):
            header_row = i
            break

        if re.search(r"freq(uency)?\s*\[", line):
            header_row = i
            break

    if header_row is None:
        # =========================================
        # FALLBACK: COBA LS TELECOM RAW AUTO-DETECT
        # =========================================
        try:
            df = pd.read_csv(
                filepath,
                sep=";",
                skiprows=11,
                encoding="latin1",
                engine="python"
            )
    
            # Ciri kuat LS Telecom: kolom frekuensi banyak & numerik
            if df.shape[1] > 10:
                df = df.iloc[:, 3:]  # buang metadata
                df_t = df.T.reset_index()
                df_t.rename(columns={"index": "Frequency (Hz)"}, inplace=True)
    
                df_t["Frequency (Hz)"] = (
                    df_t["Frequency (Hz)"]
                    .astype(str)
                    .str.replace(",", ".", regex=False)
                    .str.replace(r"[^0-9\.]", "", regex=True)
                )
                df_t["Frequency (Hz)"] = pd.to_numeric(
                    df_t["Frequency (Hz)"], errors="coerce"
                )
    
                for col in df_t.columns[1:]:
                    df_t[col] = pd.to_numeric(df_t[col], errors="coerce")
    
                df_t["Level (dBµV/m)"] = df_t.iloc[:, 1:].max(axis=1)
    
                df_final = df_t[["Frequency (Hz)", "Level (dBµV/m)"]].dropna()
    
                if not df_final.empty:
                    return df_final
    
        except Exception:
            pass  # lanjut error asli
    
        # Jika semua gagal
        raise ValueError("Header spectrum (Frequency) tidak ditemukan dan bukan LS TELECOM RAW")


    # -----------------------------
    # Gabungkan data dari header
    # -----------------------------
    data_text = "".join(lines[header_row:])
    header_line = lines[header_row]

    if ";" in header_line:
        sep = ";"
    elif "\t" in header_line:
        sep = "\t"

    df = pd.read_csv(
        StringIO(data_text),
        sep=sep,
        engine="python",
        on_bad_lines="skip"
    )

    df.columns = df.columns.astype(str).str.strip()

    # -----------------------------
    # Deteksi kolom Frequency
    # -----------------------------
    freq_cols = [
        c for c in df.columns
        if re.search(r"freq|frequency", c, re.I)
    ]
    if not freq_cols:
        raise ValueError(f"Kolom Frequency tidak ditemukan: {list(df.columns)}")
    fcol = freq_cols[0]

    # -----------------------------
    # Deteksi kolom Level
    # -----------------------------
    level_cols = [
        c for c in df.columns
        if re.search(r"level|field|strength|amplitude", c, re.I)
    ]
    if not level_cols:
        raise ValueError(f"Kolom Level tidak ditemukan: {list(df.columns)}")
    lcol = level_cols[0]

    # -----------------------------
    # Parsing numerik aman
    # -----------------------------
    freq = (
        df[fcol]
        .astype(str)
        .str.replace(",", ".", regex=False)
        .str.replace(r"[^0-9\.]", "", regex=True)
    )
    freq = pd.to_numeric(freq, errors="coerce")

    level = (
        df[lcol]
        .astype(str)
        .str.replace(",", ".", regex=False)
        .str.replace(r"[^0-9\.\-]", "", regex=True)
    )
    level = pd.to_numeric(level, errors="coerce")

    if "mhz" in fcol.lower():
        freq *= 1e6

    df_clean = pd.DataFrame({
        "Frequency (Hz)": freq,
        "Level (dBµV/m)": level
    }).dropna()

    if df_clean.empty:
        raise ValueError("Data spectrum kosong setelah parsing")

    return df_clean


def render_page():
    plots = []

    DISPLAY_COLS = [
        "No",
        "Band",
        "Pita Frekuensi",
        "Frekuensi",
        "Dinas",
        "Sub Service",
        "Kelas Emisi",
        "Identifikasi",
        "Legalitas",
        "Level (dBµV/m)"
    ]

    df_spec = load_spectrum(CURRENT_SPECTRUM, CURRENT_FILE_TYPE)
    df_rekap = pd.read_csv("rekap_baru.csv")

    # ===============================
    # BUILD LEVEL FROM SPECTRUM
    # ===============================
    df_rekap = attach_level_from_spectrum(df_rekap, df_spec)

    # ===============================
    # SORT & REBUILD NO (HZ BASED)
    # ===============================
    df_rekap = df_rekap.sort_values("Frekuensi").reset_index(drop=True)
    df_rekap["No"] = range(1, len(df_rekap) + 1)
    df_rekap.to_csv("rekap_baru.csv", index=False)
    
    for band, (fmin, fmax) in FREQ_RANGES.items():
    
        df_spec_band = df_spec[
            (df_spec["Frequency (Hz)"] >= fmin) &
            (df_spec["Frequency (Hz)"] <= fmax)
        ]
        if df_spec_band.empty:
            continue
    
        df_rekap_band = df_rekap[
            (df_rekap["Frekuensi"] * 1e6 >= fmin) &
            (df_rekap["Frekuensi"] * 1e6 <= fmax)
        ].copy()
    
        if df_rekap_band.empty:
            continue
    

        df_rekap_band["_freq_key"] = df_rekap_band["Frekuensi"]
    
        table_df = df_rekap_band[DISPLAY_COLS ]
    
        plots.append({
            "band": band,
            "freq": (df_spec_band["Frequency (Hz)"] / 1e6).tolist(),
            "level": df_spec_band["Level (dBµV/m)"].tolist(),
            "marker_freq": df_rekap_band["Frekuensi"].tolist(),
            "marker_level": df_rekap_band["Level (dBµV/m)"].tolist(),
            "marker_no": df_rekap_band["No"].tolist(),
            "table": table_df.to_dict("records")
        })

    df_rekap = pd.read_csv("rekap_baru.csv")
    summary = build_summary(df_rekap)
    dataset_name = session.get("dataset_name")


    return render_template_string(
        TEMPLATE,
        plots=plots,
        summary=summary,
        dataset_name=dataset_name
    )


UPLOAD_REGISTRY = os.path.join(UPLOAD_FOLDER, "uploads.json")

def load_uploads():
    if not os.path.exists(UPLOAD_REGISTRY):
        return []
    with open(UPLOAD_REGISTRY, "r", encoding="utf-8") as f:
        return json.load(f)

def save_uploads(data):
    with open(UPLOAD_REGISTRY, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)

@app.route("/open/<upload_id>")
def open_upload(upload_id):
    global CURRENT_SPECTRUM, CURRENT_FILE_TYPE

    uploads = load_uploads()
    u = next((x for x in uploads if x["id"] == upload_id), None)

    if not u:
        return "Data tidak ditemukan", 404

    CURRENT_SPECTRUM = os.path.join(UPLOAD_FOLDER, u["file"])
    CURRENT_FILE_TYPE = u["file_type"]

    return redirect(url_for("dashboard"))

@app.route("/identifikasi", methods=["GET", "POST"])
def identifikasi():
    global CURRENT_SPECTRUM, CURRENT_FILE_TYPE, CURRENT_DATASET_NAME

    uploads = load_uploads()

    # ======================================================
    # GET → tampilkan halaman upload
    # ======================================================
    if request.method == "GET":
        df = pd.read_csv("rekap_2025.csv")

        catatan_list = (
            df["Catatan"]
            .dropna()
            .astype(str)
            .unique()
            .tolist()
        )

        return render_template_string(
            UPLOAD_PAGE,
            catatan=sorted(catatan_list),
            uploads=uploads
        )

    # ======================================================
    # POST → proses identifikasi
    # ======================================================
    spectrum = request.files["spectrum"]
    dataset_name = request.form.get("dataset_name")
    file_type = request.form.get("file_type")
    catatan = request.form.get("catatan")

    if not dataset_name:
        return "Nama dataset wajib diisi", 400

    session["dataset_name"] = dataset_name
    
    # ===============================
    # SIMPAN FILE SPECTRUM
    # ===============================
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"spectrum_{ts}.csv"
    sp_path = os.path.join(UPLOAD_FOLDER, filename)
    spectrum.save(sp_path)

    # ===============================
    # SIMPAN REGISTRY UPLOAD
    # ===============================
    uploads.append({
        "id": ts,
        "name": dataset_name,
        "file": filename,
        "file_type": file_type,
        "catatan": catatan
    })
    save_uploads(uploads)

    # ===============================
    # SET DATA AKTIF
    # ===============================
    CURRENT_SPECTRUM = sp_path
    CURRENT_FILE_TYPE = file_type

    # ===============================
    # LOAD & FILTER REKAP
    # ===============================
    df_rekap = pd.read_csv("rekap_2025.csv")

    if catatan:
        df_rekap = df_rekap[df_rekap["Catatan"] == catatan]

    # simpan sebagai rekap aktif
    df_rekap.to_csv("rekap_baru.csv", index=False)

    # ===============================
    # LANJUT KE DASHBOARD
    # ===============================
    return redirect(url_for("dashboard"))



@app.route("/delete/<int:no>")
def delete_row(no):
    df = pd.read_csv("rekap_baru.csv")

    df = df[df["No"] != no]

    df = df.sort_values("Frekuensi").reset_index(drop=True)
    df["No"] = range(1, len(df) + 1)

    df.to_csv("rekap_baru.csv", index=False)
    return redirect(url_for("dashboard"))

@app.route("/add_click", methods=["POST"])
def add_click():
    data = request.get_json()

    freq = round(float(data["freq"]), 6)   # MHz
    level = float(data["level"])
    ident = data.get("identifikasi", "")

    df = pd.read_csv("rekap_baru.csv")

    # pastikan numeric
    df["Frekuensi"] = pd.to_numeric(df["Frekuensi"], errors="coerce")

    # ===============================
    # UPDATE JIKA FREKUENSI ADA
    # ===============================
    tol = 1e-6
    mask = abs(df["Frekuensi"] - freq) < tol

    if mask.any():
        idx = df[mask].index[0]
        df.loc[idx, "Level (dBµV/m)"] = level
        if "Identifikasi" in df.columns:
            df.loc[idx, "Identifikasi"] = ident

    # ===============================
    # INSERT JIKA BELUM ADA
    # ===============================
    else:
        ref = df[df["Frekuensi"] < freq].iloc[-1] if (df["Frekuensi"] < freq).any() else None

        new = {}
        for c in df.columns:
            if c == "No":
                continue
            elif c == "Frekuensi":
                new[c] = freq
            elif c == "Level (dBµV/m)":
                new[c] = level
            elif c.lower().startswith("identifikasi"):
                new[c] = ident
            else:
                new[c] = ref[c] if ref is not None else ""

        df = pd.concat([df, pd.DataFrame([new])], ignore_index=True)

    # ===============================
    # REBUILD NO
    # ===============================
    df = df.sort_values("Frekuensi").reset_index(drop=True)
    df["No"] = range(1, len(df) + 1)

    df.to_csv("rekap_baru.csv", index=False)
    return "", 204

@app.route("/add", methods=["POST"])
def add_row():
    fmhz = request.form.get("freq")
    if not fmhz:
        return redirect(url_for("dashboard"))

    df = pd.read_csv("rekap_baru.csv")

    hz = float(fmhz) * 1e6
    if hz in df["Frekuensi (Hz)"].values:
        return redirect(url_for("dashboard"))

    new = {c: "" for c in df.columns}
    new["Frekuensi"] = fmhz
    new["Frekuensi (Hz)"] = hz

    df = pd.concat([df, pd.DataFrame([new])], ignore_index=True)
    df = df.sort_values("Frekuensi (Hz)")
    df.to_csv("rekap_baru.csv", index=False)

    return redirect(url_for("dashboard"))

@app.route("/add_popup", methods=["POST"])
def add_popup():
    data = request.get_json()
    df = pd.read_csv("rekap_baru.csv")

    hz = float(data["Frekuensi (Hz)"])

    if hz in df["Frekuensi (Hz)"].values:
        return {"status": "exists"}, 200

    new = {}
    for col in df.columns:
        if col == "No":
            continue
        elif col == "Frekuensi (Hz)":
            new[col] = hz
        elif col == "Level (dBµV/m)":
            new[col] = float(data.get("Level (dBµV/m)", ""))
        else:
            new[col] = data.get(col, "")

    df = pd.concat([df, pd.DataFrame([new])], ignore_index=True)
    df = df.sort_values("Frekuensi (Hz)").reset_index(drop=True)
    df["No"] = range(1, len(df)+1)

    df.to_csv("rekap_baru.csv", index=False)
    return {"status": "ok"}, 200


def build_summary(df):

    summary = {}

    # ===============================
    # Legalitas
    # ===============================
    summary["legalitas"] = (
        df.groupby("Legalitas")
        .size()
        .reset_index(name="Jumlah")
        .to_dict("records")
    )

    # ===============================
    # Band
    # ===============================
    summary["band"] = (
        df.groupby(["Band"])
        .size()
        .reset_index(name="Jumlah")
        .to_dict("records")
    )

    # ===============================
    # Dinas
    # ===============================
    summary["dinas"] = (
        df.groupby(["Dinas"])
        .size()
        .reset_index(name="Jumlah")
        .to_dict("records")
    )

    # ===============================
    # Pita Frekuensi
    # ===============================
    summary["pita"] = (
        df.groupby(["Pita Frekuensi", "Legalitas"])
        .size()
        .reset_index(name="Jumlah")
        .to_dict("records")
    )

    return summary

@app.route("/save_table", methods=["POST"])
def save_table():
    data = request.get_json()
    rows = data.get("rows", {})

    df = pd.read_csv("rekap_baru.csv")

    for no, cols in rows.items():
        no = int(no)
        idx = df[df["No"] == no].index
        if idx.empty:
            continue
        idx = idx[0]

        for col, val in cols.items():
            if col in df.columns:
                df.loc[idx, col] = val

    df.to_csv("rekap_baru.csv", index=False)
    return {"status":"ok"}


@app.route("/dashboard")
def dashboard():
    return render_page()

# ======================================================
# HTML
# ======================================================
UPLOAD_PAGE = """
<!DOCTYPE html>
<html>
<head>
<title>Identifikasi Spectrum</title>

<style>
body {
    font-family: Arial, sans-serif;
    background:#f4f6f9;
    padding:40px;
}
.container {
    max-width:900px;
    margin:auto;
}
.card {
    background:white;
    padding:25px;
    border-radius:10px;
    box-shadow:0 4px 10px rgba(0,0,0,0.08);
}

label { display:block; margin-top:15px; font-weight:bold }
select, input[type=file], input[type=text] {
    width:100%;
    padding:8px;
    margin-top:6px;
}

button {
    margin-top:25px;
    background:#e1ae05;
    color:white;
    border:none;
    padding:12px 28px;
    font-size:16px;
    border-radius:6px;
    cursor:pointer;
}
button:hover { background:#c89604 }

/* ===== MODAL ===== */
.modal {
    display:none;
    position:fixed;
    inset:0;
    background:rgba(0,0,0,0.45);
}
.modal-content {
    background:white;
    width:420px;
    margin:120px auto;
    padding:25px;
    border-radius:10px;
}
.modal-content h3 { margin-top:0 }
.modal-content .note {
    font-size:13px;
    color:#555;
    background:#f8f8f8;
    padding:10px;
    border-left:4px solid #e1ae05;
    margin-bottom:15px;
}
.modal-actions {
    margin-top:20px;
    text-align:right;
}
.modal-actions button {
    margin-left:10px;
}
</style>
</head>

<body>
<div class="container">
<div class="card">

<h2>🔍 Identifikasi Spectrum</h2>
<p>Upload data pengukuran dan tentukan referensi rekap identifikasi.</p>

<form id="uploadForm" method="post" enctype="multipart/form-data">

<!-- ================== FILE PENGUKURAN ================== -->
<label>Jenis File Pengukuran</label>
<select name="file_type" required>
    <option value="">-- Pilih Jenis File --</option>
    <option value="ARGUS">ARGUS</option>
    <option value="ARGUS V6">ARGUS V6</option>
    <option value="LS TELECOM">LS TELECOM</option>
    <option value="TCI">TCI</option>
</select>

<label>File Spectrum CSV</label>
<input type="file" name="spectrum" accept=".csv" required>

<!-- ================== REFERENSI ================== -->
<h3 style="margin-top:30px">📌 Tentukan Referensi</h3>
<p style="font-size:13px;color:#555">
Data diambil dari <b>Rekapitulasi ROL 2025</b>
</p>

<label>Catatan</label>
<select name="catatan">
    <option value="">-- Semua Catatan --</option>
    {% for v in catatan %}
    <option value="{{ v }}">{{ v }}</option>
    {% endfor %}
</select>

<!-- hidden field -->
<input type="hidden" name="dataset_name" id="dataset_name">

<button type="button" onclick="openPopup()">Proses Identifikasi</button>


<!-- ================== LIST DATASET ================== -->
<h3 style="margin-top:40px">📂 Spectrum Tersimpan</h3>
<ul class="upload-list">
{% for u in uploads %}
<li>
<a href="/open/{{ u.id }}">📊 {{ u.name }}</a>
</li>
{% else %}
<li><i>Belum ada data</i></li>
{% endfor %}
</ul>

</div>
</div>

</form>

<!-- ================== POPUP NAMA DATASET ================== -->
<div id="popup" class="modal">
<div class="modal-content">
<h3>📝 KASIH NAMA DULU CUYYY!!</h3>

<div class="note">
<b>Aturan penamaan:</b><br>
Kegiatan – Site – Bulan Tahun<br>
<i>Contoh:</i><br>
Identifikasi - TCI Bukit Bakan - Januari 2026
</div>


<label>Nama Penyimpanan</label>
<input type="text" id="dataset_input" placeholder="Identifikasi TCI Bukit Bakan Januari 2026">

<div class="modal-actions">
<button onclick="closePopup()">Batal</button>
<button onclick="submitUpload()">Simpan & Proses</button>

</div>

</div>
</div>

</div>
</div>



<script>
function openPopup(){
    document.getElementById("popup").style.display = "block";
}

function closePopup(){
    document.getElementById("popup").style.display = "none";
}

function submitUpload(){
    const name = document.getElementById("dataset_input").value.trim();
    if(!name){
        alert("Nama dataset wajib diisi");
        return;
    }
    document.getElementById("dataset_name").value = name;
    document.getElementById("uploadForm").submit();
}
</script>

</body>
</html>
"""


TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
<title>Spectrum Dashboard</title>
<script src="https://cdn.plot.ly/plotly-2.27.0.min.js"></script>
<style>
.rekap-grid {
    display:grid;
    grid-template-columns: repeat(auto-fit,minmax(320px,1fr));
    gap:20px;
    margin-bottom:40px;
}
.rekap-card {
    background:#ffffff;
    padding:15px;
    border-radius:10px;
    box-shadow:0 2px 6px rgba(0,0,0,.08);
}
.rekap-card h3 {
    margin-top:0;
    font-size:15px;
}
.rekap-card table {
    width:100%;
    border-collapse:collapse;
    font-size:12px;
}
.rekap-card th,
.rekap-card td {
    border:1px solid #ccc;
    padding:6px;
    text-align:left;
}
.rekap-card th {
    background:#f0f0f0;
}

:root{
    --primary:#1f4fd8;
    --danger:#d9534f;
    --bg:#f4f6f9;
    --card:#ffffff;
    --border:#dcdcdc;
}

body {
    font-family: "Segoe UI", Arial, sans-serif;
    background: var(--bg);
    padding: 30px;
    color:#333;
}

h2 {
    margin-bottom: 30px;
    color:#1a1a1a;
}

.card {
    background: var(--card);
    padding: 22px;
    margin-bottom: 40px;
    border-radius: 14px;
    box-shadow: 0 8px 20px rgba(0,0,0,0.08);
}

.card h3 {
    margin-top: 0;
    color: var(--primary);
}

/* ================= TABLE ================= */
table {
    border-collapse: collapse;
    width:100%;
    margin-top:20px;
    font-size:13px;
}

th {
    background:#eef2f7;
    padding:10px 8px;
    border:1px solid var(--border);
    font-weight:600;
}

td {
    border:1px solid var(--border);
    padding:8px 6px;
    text-align:center;
}

tr:nth-child(even) {
    background:#fafafa;
}

td[contenteditable="true"] {
    background:#fffef5;
    cursor:text;
}

td[contenteditable="true"]:focus {
    outline:none;
    background:#fff8dc;
    box-shadow: inset 0 0 0 1px var(--primary);
}

/* ================= BUTTON ================= */
a.btn-delete {
    color: white;
    background: var(--danger);
    padding:4px 10px;
    border-radius:6px;
    text-decoration:none;
    font-size:12px;
}

a.btn-delete:hover {
    opacity:0.85;
}

/* ================= MODAL ================= */
.modal {
    display:none;
    position:fixed;
    inset:0;
    background:rgba(0,0,0,0.45);
    z-index:1000;
}

.modal-content {
    background:white;
    width:380px;
    margin:120px auto;
    padding:22px;
    border-radius:14px;
    box-shadow:0 12px 30px rgba(0,0,0,0.25);
}

.modal-content h4 {
    margin-top:0;
    color:var(--primary);
}

.modal-content label {
    display:block;
    margin-top:12px;
    font-size:13px;
    font-weight:600;
}

.modal-content input {
    width:100%;
    padding:8px;
    margin-top:5px;
    border-radius:6px;
    border:1px solid var(--border);
}

.modal-content button {
    margin-top:18px;
    padding:8px 16px;
    border:none;
    border-radius:8px;
    cursor:pointer;
    font-size:14px;
}

.modal-content button:first-of-type {
    background:var(--primary);
    color:white;
}

.modal-content button:last-of-type {
    background:#ccc;
    margin-left:8px;
}
.subtitle {
    margin-top: -10px;
    margin-bottom: 20px;
    color: #555;
    font-size: 14px;
}
.save-btn {
    background:#1e88e5;
    color:white;
    border:none;
    padding:6px 14px;
    font-size:13px;
    border-radius:6px;
    cursor:pointer;
}
.save-btn:hover {
    background:#1565c0;
}


</style>

</head>

<body>
<h2>
Spectrum Monitoring
{% if dataset_name %}
<br>
<small style="color:#555">📁 {{ dataset_name }}</small>
{% endif %}
</h2>



<div class="card">
<h3>📊 Rekapitulasi Identifikasi</h3>

<!-- ================= LEGALITAS ================= -->
<h4>Rangkuman Berdasarkan Legalitas</h4>
<table>
<tr><th>Legalitas</th><th>Jumlah</th></tr>
{% for r in summary.legalitas %}
<tr>
<td>{{ r.Legalitas }}</td>
<td>{{ r.Jumlah }}</td>
</tr>
{% endfor %}
</table>

<!-- ================= BAND ================= -->
<h4>Rangkuman Berdasarkan Band</h4>
<table>
<tr><th>Band</th><th>Jumlah</th></tr>
{% for r in summary.band %}
<tr>
<td>{{ r.Band }}</td>
<td>{{ r.Jumlah }}</td>
</tr>
{% endfor %}
</table>

<!-- ================= DINAS ================= -->
<h4>Rangkuman Berdasarkan Dinas</h4>
<table>
<tr><th>Dinas</th><th>Jumlah</th></tr>
{% for r in summary.dinas %}
<tr>
<td>{{ r.Dinas }}</td>
<td>{{ r.Jumlah }}</td>
</tr>
{% endfor %}
</table>

<!-- ================= PITA ================= -->
<h4>Rangkuman Berdasarkan Pita Frekuensi</h4>
<table>
<tr><th>Pita Frekuensi</th><th>Legalitas</th><th>Jumlah</th></tr>
{% for r in summary.pita %}
<tr>
<td>{{ r["Pita Frekuensi"] }}</td>
<td>{{ r.Legalitas }}</td>
<td>{{ r.Jumlah }}</td>
</tr>
{% endfor %}
</table>
</div>


{% for p in plots %}
<div class="card">

<h3>{{ p.band }}</h3>
<div id="plot-{{ loop.index }}" style="height:360px;"></div>

<div style="text-align:right; margin-top:10px">
    <button class="save-btn"
        onclick="saveTable({{ loop.index }})">
        💾 Simpan Perubahan
    </button>
</div>

<table>
<tr>
{% for k in p.table[0].keys() %}
<th>{{ k }}</th>
{% endfor %}
<th>Aksi</th>
</tr>

{% for r in p.table %}
<tr>
{% for k,v in r.items() %}
<td contenteditable="true"
    data-row="{{ r['No'] }}"
    data-col="{{ k }}">
    {{ v }}
</td>

{% endfor %}
<td>
<a class="btn-delete"
   href="/delete/{{ r['No'] }}"
   onclick="return confirm('Hapus data ini?')">
   Hapus
</a>

</td>
</tr>
{% endfor %}
</table>


<!-- POPUP -->
<div id="modal-{{ loop.index }}" class="modal">
<div class="modal-content">
<h4>Tambah Data</h4>

<label>Identifikasi</label>
<input id="ident-{{ loop.index }}">

<label>Frekuensi (MHz)</label>
<input id="freq-{{ loop.index }}" readonly>

<label>Level (dBµV/m)</label>
<input id="level-{{ loop.index }}" readonly>

<button onclick="submit{{ loop.index }}()">Simpan</button>
<button onclick="close{{ loop.index }}()">Batal</button>
</div>
</div>

<script>
(function(){
var freq = {{ p.freq }};
var level = {{ p.level }};
var mFreq = {{ p.marker_freq }};
var mNo = {{ p.marker_no }};

function nearestY(f){
    let d=Infinity, idx=0;
    for(let i=0;i<freq.length;i++){
        let x=Math.abs(freq[i]-f);
        if(x<d){ d=x; idx=i; }
    }
    return level[idx];
}

var spectrum = { x:freq, y:level, mode:'lines', name:'Spectrum' };

var my = mFreq.map(nearestY);

var mc = {
    x:mFreq, y:my, mode:'markers',
    marker:{ size:12, symbol:'circle-open', line:{width:2,color:'red'} }
};

var mt = {
    x:mFreq, y:my, mode:'text',
    text:mNo, textposition:'top center'
};

Plotly.newPlot('plot-{{ loop.index }}',[spectrum,mc,mt]);

document.getElementById('plot-{{ loop.index }}')
.on('plotly_click', function(d){
    let fx = d.points[0].x;

    // Cari indeks frekuensi spectrum terdekat
    let min = Infinity, idx = 0;
    for(let i=0;i<freq.length;i++){
        let dx = Math.abs(freq[i] - fx);
        if(dx < min){
            min = dx;
            idx = i;
        }
    }

    let f = freq[idx];
    let y = level[idx];

    document.getElementById('freq-{{ loop.index }}').value = f.toFixed(6);
    document.getElementById('level-{{ loop.index }}').value = y.toFixed(2);
    document.getElementById('modal-{{ loop.index }}').style.display='block';
});


window.close{{ loop.index }} = ()=> {
    document.getElementById('modal-{{ loop.index }}').style.display='none';
}

window.submit{{ loop.index }} = ()=> {
    fetch('/add_click',{
        method:'POST',
        headers:{'Content-Type':'application/json'},
        body:JSON.stringify({
            identifikasi: document.getElementById('ident-{{ loop.index }}').value,
            freq: document.getElementById('freq-{{ loop.index }}').value,
            level: document.getElementById('level-{{ loop.index }}').value
        })
    }).then(()=>location.reload());
}
})();
</script>
<script>
function saveTable(idx){
    const table = document.querySelectorAll(
        '#plot-' + idx
    )[0].parentElement.querySelectorAll("td[data-row]");

    let rows = {};

    table.forEach(td => {
        const r = td.dataset.row;
        const c = td.dataset.col;
        const v = td.innerText.trim();

        if(!rows[r]) rows[r] = {};
        rows[r][c] = v;
    });

    fetch("/save_table",{
        method:"POST",
        headers:{ "Content-Type":"application/json" },
        body: JSON.stringify({ rows: rows })
    })
    .then(r => r.json())
    .then(res => {
        alert("Perubahan berhasil disimpan");
    })
    .catch(err=>{
        alert("Gagal menyimpan data");
    });
}
</script>
</div>
{% endfor %}
</body>
</html>
"""


# Middleware untuk proteksi halaman
def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "username" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function

# Ambil data dari API / Google Sheet / CSV
def load_map_data():
    sheet_id = "14rFJPrA2fCVkz-7mQoLQ8khV5nLlsKVv"
    gid = "301785538"

    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"

    # Ambil data ke DataFrame
    df_map = pd.read_csv(url, header=1)
    return df_map

def generate_map_html_from_df(df_map, out_filename="map_folium.html"):
    lat_col = 'Latitude'
    lon_col = 'Longitude'

    if lat_col not in df_map.columns or lon_col not in df_map.columns:
        return None
    
    # Konversi ke float
    for col in [lat_col, lon_col]:
        df_map[col] = (
            df_map[col].astype(str).str.replace(",", ".", regex=False)
                       .astype(float)
        )

    # Hapus baris invalid
    df_valid = df_map.dropna(subset=[lat_col, lon_col])
    
    if df_valid.empty:
        return None

    # Tengah peta
    center_lat = float(df_valid[lat_col].mean())
    center_lon = float(df_valid[lon_col].mean())

    m = folium.Map(location=[center_lat, center_lon], zoom_start=8)
    marker_cluster = MarkerCluster().add_to(m)
    
    for i, row in df_valid.iterrows():

        # Ambil nilai Speedtest
        dl = row.get("Average Speedtest Download Speed (Mbps)", "N/A")
        ul = row.get("Average Speedtest Upload Speed (Mbps)", "N/A")

        # Ambil nama kab/kota
        kabkot = row.get("Kabupaten / Kota", "Tidak diketahui")

        # Popup HTML rapi
        popup_text = f"""
        <b>{kabkot}</b><br>
        <hr style='margin:4px 0;'>
        <b>📥 Download:</b> {dl} Mbps<br>
        <b>📤 Upload:</b> {ul} Mbps<br>
        """

        folium.Marker(
            location=[row[lat_col], row[lon_col]],
            popup=folium.Popup(popup_text, max_width=350)
        ).add_to(marker_cluster)

    # Simpan map
    os.makedirs("static", exist_ok=True)
    file_path = os.path.join("static", out_filename)
    m.save(file_path)

    return out_filename


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")

        if username in USERS and USERS[username] == password:
            session["username"] = username
            flash("Login berhasil!", "success")
            return redirect(url_for("index"))
        else:
            flash("Username atau password salah!", "danger")

    return render_template_string("""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Login WANDAA</title>
        <style>
            body {
                margin: 0;
                padding: 0;
                font-family: Arial, sans-serif;
                background: url('{{ url_for("static", filename="kolase.png") }}') no-repeat center center fixed;
                background-size: cover;
                display: flex;
                justify-content: center;
                align-items: center;
                height: 100vh;
                color: white;
            }
            .login-box {
                background: rgba(0, 0, 0, 0.7);
                padding: 30px;
                border-radius: 12px;
                width: 320px;
                text-align: center;
                box-shadow: 0 4px 10px rgba(0,0,0,0.5);
            }
            .login-box h2 {
                margin-bottom: 20px;
                font-size: 22px;
            }
            .login-box input {
                width: 100%;
                padding: 10px;
                margin: 10px 0;
                border: none;
                border-radius: 6px;
                font-size: 14px;
            }
            .login-box button {
                width: 100%;
                padding: 10px;
                background: #006db0;
                border: none;
                border-radius: 6px;
                color: white;
                font-weight: bold;
                cursor: pointer;
                transition: background 0.3s;
            }
            .login-box button:hover {
                background: #004d80;
            }
            .flash-message {
                margin-top: 10px;
                color: #ff6b6b;
                font-size: 14px;
            }
        </style>
    </head>
    <body>
        <div class="login-box">
            <h2>ONE-DATA AGGREGATION AND ANALYTICS (WANDAA)</h2>
            <form method="POST">
                <input type="text" name="username" placeholder="Username" required>
                <input type="password" name="password" placeholder="Password" required>
                <button type="submit">Login</button>
            </form>
            {% with messages = get_flashed_messages(with_categories=true) %}
              {% if messages %}
                <div class="flash-message">
                {% for category, message in messages %}
                  <p>{{ message }}</p>
                {% endfor %}
                </div>
              {% endif %}
            {% endwith %}
        </div>
    </body>
    </html>
    """)


@app.route("/logout")
def logout():
    session.pop("username", None)
    return redirect(url_for("login"))
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

logging.basicConfig(
    filename=os.path.join(BASE_DIR, "app_log.txt"),
    level=logging.ERROR,
    format="%(asctime)s [%(levelname)s] %(message)s"
)

# Bersihkan string agar aman untuk nama file
def safe_filename(text):
    return re.sub(r'[\\/*?:"<>|]', "_", str(text).strip())

def load_data(year):
    url = "https://rol.postel.go.id/api/observasi/allapproved"
    params = {
        "upt": 19,
        "year": year,
        "pageIndex": 1,
        "pageSize": 100000
    }
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
        "Accept": "*/*",
        "Referer": "https://rol.postel.go.id/observasi/laporan",
        "X-Requested-With": "XMLHttpRequest",
        "Cookie": "csrf_cookie_name=6701fbfe229f77d47c710eec3391f386; ci_session=1tmigk58v1636foa82v97rmo8o2olsrc"
    }

    try:
        response = requests.get(url, params=params, headers=headers)
        data = response.json().get("data", [])
        return pd.DataFrame(data)
    except Exception as e:
        print("Gagal ambil data API:", e)
        return pd.DataFrame()

def load_info_inspeksi(use_cache=True):
    url = "https://apstard.postel.go.id/dashboard/info-inspeksi-3"

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/139.0.0.0 Safari/537.36 Edg/139.0.0.0",
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": "https://apstard.postel.go.id/dashboard/dashboard-keseluruhan-upt",
        # ⚠️ Cookie perlu diganti sesuai hasil login
        "Cookie": "csrf_cookie_name=ISI_COOKIE; ci_session=ISI_SESSION",
    }

    payload = {
        "periode": "2025",
        "upt_id": "14"
    }

    try:
        r = requests.post(url, headers=headers, data=payload, timeout=10)
        r.raise_for_status()

        data = r.json()

        # Simpan ke cache lokal
        if use_cache:
            with open("inspeksi.json", "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

        return pd.DataFrame([data]), None  # DataFrame + status OK

    except Exception as e:
        print(f"⚠️ Gagal ambil data dari Apstard: {e}")

        # Gunakan cache lokal kalau ada
        if use_cache and os.path.exists("inspeksi.json"):
            print("👉 Memuat data dari cache lokal inspeksi.json")
            with open("inspeksi.json", "r", encoding="utf-8") as f:
                data = json.load(f)
            return pd.DataFrame([data]), "⚠️ Data dari cache (Apstard tidak bisa diakses)"

        # Kalau tidak ada cache
        return pd.DataFrame(), "⚠️ Aplikasi Apstard tidak bisa diakses"

def get_invoice_session():
    """
    Membuat session login invoice dengan auto-cookie
    """
    sess = requests.Session()
    sess.headers.update({
        "User-Agent": "Mozilla/5.0",
        "Accept": "application/json",
        "X-Requested-With": "XMLHttpRequest"
    })

    payload = {
        "username": INVOICE_USERNAME,
        "password": INVOICE_PASSWORD
    }

    r = sess.post(INVOICE_LOGIN_URL, json=payload, timeout=15)
    r.raise_for_status()

    return sess

def load_invoice_data(params=None, use_cache=True):
    """
    Ambil data invoice dengan auto relogin jika token expired
    """
    cache_file = "invoice.json"

    try:
        sess = get_invoice_session()
        r = sess.post(INVOICE_DATA_URL, json=params or {}, timeout=20)

        # Jika server error / token mati
        if r.status_code in [401, 403, 500]:
            raise Exception("Session expired")

        data = r.json()

        if use_cache:
            with open(cache_file, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2)

        return pd.DataFrame(data.get("data", []))

    except Exception as e:
        print("⚠️ Gagal ambil invoice:", e)

        # fallback cache
        if use_cache and os.path.exists(cache_file):
            print("👉 Pakai cache invoice.json")
            with open(cache_file, "r", encoding="utf-8") as f:
                data = json.load(f)
            return pd.DataFrame(data.get("data", []))

        return pd.DataFrame()


def load_pantib(year, use_cache=True):
    url = "https://rol.postel.go.id/api/penertiban/list"

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/139.0.0.0 Safari/537.36 Edg/139.0.0.0",
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": "https://rol.postel.go.id/penertiban",
        # ⚠️ Cookie selalu berubah → kalau tidak valid, fallback ke pantib.json
        "Cookie": "csrf_cookie_name=2e099f4d4f3f335b4f9b445017d17869; ci_session=i2rgrkn5ha876ru0fnm6i9pqmqtg1bho"
    }

    session = requests.Session()
    retries = Retry(connect=3, backoff_factor=1, status_forcelist=[500, 502, 503, 504])
    session.mount("https://", HTTPAdapter(max_retries=retries))

    all_data = []
    page = 1
    page_size = 1000  

    try:
        while True:
            params = {
                "status": "",
                "status_penertiban": "",
                "tahun": year,
                "pageIndex": page,
                "pageSize": page_size
            }
            r = session.get(url, headers=headers, params=params, timeout=15)

            # kalau cookie invalid → biasanya balasan HTML login page
            if "<title>ROL Login</title>" in r.text:
                raise Exception("Cookie expired / invalid")

            data = r.json().get("data", [])
            if not data:
                break
            all_data.extend(data)
            page += 1

        # simpan cache lokal
        if use_cache:
            with open("pantib.json", "w", encoding="utf-8") as f:
                json.dump(all_data, f, ensure_ascii=False, indent=2)

    except Exception as e:
        print("⚠️ Gagal ambil data dari API pantib:", e)

        # fallback ke cache lokal
        if use_cache and os.path.exists("pantib.json"):
            print("👉 Memuat data dari cache lokal pantib.json")
            with open("pantib.json", "r", encoding="utf-8") as f:
                all_data = json.load(f)
        else:
            print("⚠️ Tidak ada cache pantib.json ditemukan")

    return pd.DataFrame(all_data)



def format_tanggal_indonesia(tgl_raw):
    bulan_map = {
        "January": "Januari", "February": "Februari", "March": "Maret",
        "April": "April", "May": "Mei", "June": "Juni",
        "July": "Juli", "August": "Agustus", "September": "September",
        "October": "Oktober", "November": "November", "December": "Desember"
    }

    try:
        # parse string YYYY-MM-DD
        dt = datetime.strptime(tgl_raw, "%Y-%m-%d")
        # cek OS (Windows pakai %#d, Linux pakai %-d)
        try:
            day_str = dt.strftime("%-d")   # Linux/Unix
        except:
            day_str = dt.strftime("%#d")  # Windows
        month_str = dt.strftime("%B")
        year_str = dt.strftime("%Y")

        # ganti bulan ke bahasa Indonesia
        month_str = bulan_map.get(month_str, month_str)

        return f"{day_str} {month_str} {year_str}"
    except Exception:
        return tgl_raw  # fallback kalau parsing gagal


@app.route("/unduh_laporan", methods=["GET", "POST"])
def unduh_laporan():     
    selected_year = request.form.get("year", "2026")
    df = load_data(selected_year)

    # Transformasi jenis identifikasi
    df["observasi_status_identifikasi_name"] = df["observasi_status_identifikasi_name"].str.replace(
        r"OFF AIR \(Sedang Tidak Digunakan\)", "OFF AIR", regex=True
    )
    df['jenis'] = df['observasi_status_identifikasi_name'].apply(
        lambda x: 'Belum Teridentifikasi' if x == 'BELUM DIKETAHUI' else 'Teridentifikasi'
    )
    
    locale.setlocale(locale.LC_TIME, "id_ID.utf8")  # aktifkan format tanggal bahasa Indonesia
    tanggal_skrg = datetime.now().strftime("%d %B %Y")

    # Ambil filter dari form (POST) atau query (GET)
    if request.method == "POST":
        selected_spt = request.form.get("spt", "Semua")
        selected_kab = request.form.get("kab", "Semua")
        selected_kec = request.form.get("kec", "Semua")
        selected_cat = request.form.get("cat", "Semua")
        pelaksana_list = request.form.getlist("pelaksana")
        perangkat = request.form.get("perangkat", "Tetap/Transportable TCI")
        tgl_spt_raw = request.form.get("tgl_spt")
        if tgl_spt_raw:
                tgl_spt = format_tanggal_indonesia(tgl_spt_raw)
        else:
                tgl_spt = format_tanggal_indonesia(datetime.now().strftime("%Y-%m-%d"))
        
    else:
        selected_spt = request.args.get("spt", "Semua")
        selected_kab = request.args.get("kab", "Semua")
        selected_kec = request.args.get("kec", "Semua")
        selected_cat = request.args.get("cat", "Semua")
        pelaksana_list = []
        tgl_spt = format_tanggal_indonesia(datetime.now().strftime("%Y-%m-%d"))
        perangkat = "Tetap/Transportable TCI"

    # Filter data
    filt = df.copy()
    if selected_spt != "Semua":
        filt = filt[filt["observasi_no_spt"] == selected_spt]
    if selected_kab != "Semua":
        filt = filt[filt["observasi_kota_nama"] == selected_kab]
    if selected_kec != "Semua":
        filt = filt[filt["observasi_kecamatan_nama"] == selected_kec]
    if selected_cat != "Semua":
        filt = filt[filt["scan_catatan"] == selected_cat]

    # pastikan kolom tanggal dalam bentuk datetime
    filt["observasi_tanggal"] = pd.to_datetime(filt["observasi_tanggal"], errors="coerce")
    
    # ambil bulan & tahun dari data observasi
    if not filt.empty and filt["observasi_tanggal"].notna().any():
        bulan_tahun_obs = filt["observasi_tanggal"].dt.strftime("%B %Y").iloc[0]
    else:
        bulan_tahun_obs = datetime.now().strftime("%B %Y")


    # === Registrasi font ===
    pdfmetrics.registerFont(UnicodeCIDFont('HeiseiMin-W3'))
    #pdfmetrics.registerFont(TTFont("BrushScript", "BRUSHSCI.ttf"))
    pdfmetrics.registerFont(TTFont("BrushScript", "BRUSHSCI.ttf"))
    pdfmetrics.registerFont(TTFont("zph", "bodoni-six-itc-bold-italic-os-5871d33e4dc4a.ttf"))
    pdfmetrics.registerFont(TTFont("Arial", "ARIALBD.ttf"))
    pdfmetrics.registerFont(TTFont("Arialbd", "ARIAL.ttf"))

    # Simpan sementara
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    filename = tmp.name
    doc = SimpleDocTemplate(filename, pagesize=A4,
                            rightMargin=50, leftMargin=50,
                            topMargin=20, bottomMargin=30)

    styles = getSampleStyleSheet()
    style_center = ParagraphStyle(name="Center", parent=styles["Normal"], alignment=TA_CENTER, fontName="Arial", fontSize=14, leading=15)
    style_center2 = ParagraphStyle(name="Center", parent=styles["Normal"], alignment=TA_CENTER, fontName="Arialbd", fontSize=12, leading=15)
    style_normal = ParagraphStyle(name="Normal", parent=styles["Normal"], alignment=TA_JUSTIFY, fontName="Arialbd", fontSize=12, leading=14)

    style_left_h1b = ParagraphStyle(name="Left", parent=styles["Normal"], alignment=TA_LEFT, fontName="zph", fontSize=16, leading=16, textColor=colors.blue)
    style_left_h2b = ParagraphStyle(name="Left", parent=styles["Normal"], alignment=TA_LEFT, fontName="BrushScript", fontSize=14, leading=16, textColor=colors.blue)
    style_left_h3b = ParagraphStyle(name="Left", parent=styles["Normal"], alignment=TA_LEFT, fontName="zph", fontSize=11, leading=16, textColor=colors.blue)
    style_left_h4b = ParagraphStyle(name="Left", parent=styles["Normal"], alignment=TA_LEFT, fontName="zph", fontSize=9, leading=16, textColor=colors.blue)
    style_left_h4t = ParagraphStyle(name="Left", parent=styles["Normal"], alignment=TA_LEFT, fontName="Arialbd", fontSize=12, leading=14)

    content = []
    
    # === Kop Surat ===
    base_dir = os.path.dirname(os.path.abspath(__file__))  # folder tempat file .py berada
    logo_path = os.path.join(base_dir, "static", "logo-kominfo.png")
    
    logo = Image(logo_path, width=70, height=70)
    #logo = "LOGO KOMINFO"
    
    kop_text = [
        Paragraph("<b>KEMENTERIAN KOMUNIKASI DAN INFORMATIKA RI</b>", style_left_h1b),
        Paragraph("DIREKTORAT JENDERAL SUMBER DAYA DAN PERANGKAT POS DAN INFORMATIKA", style_left_h3b),
        Paragraph("BALAI MONITOR SPEKTRUM FREKUENSI RADIO KELAS II MATARAM", style_left_h3b),
        Paragraph("Indonesia Terkoneksi : Makin Digital, Makin Maju", style_left_h2b),
        Paragraph("Jl.Singosari No.4 Mataram 83127 Telp.(0370) 646411 Fax.(0370) 648740-42 email: upt_mataram.postel.go.id", style_left_h4b)
    ]
    kop_table = Table([[logo, kop_text]], colWidths=[70, 430])
    kop_table.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("LINEBELOW", (0,0), (-1,-1), 2, colors.blue)
    ]))
    content.append(kop_table)
    content.append(Spacer(1, 20))        
        
    # === Judul ===
    content.append(Paragraph("<b>NOTA DINAS</b>", style_center))
    # ambil bulan/tahun sekarang
    bulan_tahun = datetime.now().strftime("%m/%Y")
    content.append(Paragraph(f"<b>Nomor :&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;/ND/Montib/{bulan_tahun}</b>", style_center2))
    content.append(Spacer(1, 20))

    # === Tabel Yth ===
    yth_text = [
        Paragraph("Yth", style_left_h4t),
        Paragraph("Dari", style_left_h4t),
        Paragraph("Hal", style_left_h4t),
        Paragraph("Sifat", style_left_h4t),
        Paragraph("Lampiran", style_left_h4t),
        Paragraph("Tanggal", style_left_h4t)
    ]
    yth2_text = [
        Paragraph("Kepala Balai Monitor SFR Kelas II Mataram", style_normal),
        Paragraph("Ketua Tim Kerja Monitoring dan Penertiban SFR dan APT", style_normal),
        Paragraph(f"Laporan Pelaksanaan Kegiatan Monitoring dan Identifikasi 15 Pita Frekuensi Radio dengan Perangkat SMFR {perangkat} Site {selected_kec}, {selected_kab} Bulan Agustus Tahun 2025", style_normal),
        Paragraph("Biasa", style_left_h4t),
        Paragraph("Satu bendel", style_left_h4t),
        Paragraph(f"{tanggal_skrg}", style_left_h4t)
    ]
    data = [[l, ":", r] for l, r in zip(yth_text, yth2_text)]
    yth_table = Table(data, colWidths=[70, 10, 420])
    yth_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),   # rata atas
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),   # tetap rata kiri
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 11),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    content.append(yth_table)
    content.append(Spacer(1, 20))
    
    # === Isi ===
    isi = f"""
    Dengan hormat disampaikan, bahwa berdasarkan Surat Tugas Nomor : 
    {selected_spt}, tanggal {tgl_spt}, tentang Kegiatan 
    Monitoring/Observasi dan Identifikasi 15 Pita Frekuensi Radio dengan Pemanfaatan 
    Perangkat SMFR {perangkat} Bulan {bulan_tahun_obs}, 
    terlampir kami sampaikan laporan pelaksanaan kegiatan dimaksud untuk 
    SMFR {perangkat} Site {selected_kec}, {selected_kab}.<br/><br/>
    
    Demikian disampaikan, mohon arahan lebih lanjut dan atas perhatian Bapak 
    diucapkan terimakasih.
    """
    content.append(Paragraph(isi, style_normal))
    content.append(Spacer(1, 80))
    

        
    # === TTD ===
    style_center_block = ParagraphStyle(name="CenterBlock", parent=styles["Normal"], alignment=TA_CENTER, fontName="Arialbd", fontSize=12)
    
    pelaksana_list = request.form.getlist("pelaksana")
    
    pelaksana_table = Table(
        [["", Paragraph("Abdy Budiman Djara", style_center_block)]],
        colWidths=[350,150]  # sesuaikan lebar halaman
    )
    pelaksana_table.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "RIGHT")
    ]))
    content.append(pelaksana_table)
    content.append(Spacer(1, 50))        
    
    # Tambah pemisah halaman
    content.append(PageBreak())
    
    # === Kop Surat ===
    content.append(kop_table)
    content.append(Spacer(1, 20))        
        
    # === Judul ===
    content.append(Paragraph("<b>NOTA DINAS</b>", style_center))
    content.append(Spacer(1, 20))

    # === Tabel Yth ===
    yth_text = [
        Paragraph("Yth", style_left_h4t),
        Paragraph("Dari", style_left_h4t),
        Paragraph("Hal", style_left_h4t),
        Paragraph("Sifat", style_left_h4t),
        Paragraph("Tanggal", style_left_h4t)
    ]
    yth2_text = [
        Paragraph("Ketua Tim Kerja Monitoring dan Penertiban SFR dan APT", style_normal),
        Paragraph(f"Pelaksana Kegiatan Monitoring dan Identifikasi 15 Pita Frekuensi Radio dengan Perangkat SMFR {perangkat} Site {selected_kec}, {selected_kab}", style_normal),
        Paragraph(f"Laporan Pelaksanaan Kegiatan Monitoring dan Identifikasi 15 Pita Frekuensi Radio dengan Perangkat SMFR {perangkat} Site {selected_kec}, {selected_kab} Bulan Agustus Tahun 2025", style_normal),
        Paragraph("Biasa", style_left_h4t),
        Paragraph(f"{tanggal_skrg}", style_left_h4t)
    ]
    data = [[l, ":", r] for l, r in zip(yth_text, yth2_text)]
    yth_table = Table(data, colWidths=[70, 10, 420])
    yth_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),   # rata atas
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),   # tetap rata kiri
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 11),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    content.append(yth_table)
    content.append(Spacer(1, 20))
    
    # === Isi ===
    isi = f"""
    Dengan hormat disampaikan, bahwa berdasarkan Surat Tugas Nomor : 
    {selected_spt}, tanggal {tgl_spt}, tentang Kegiatan 
    Monitoring/Observasi dan Identifikasi 15 Pita Frekuensi Radio dengan Pemanfaatan 
    Perangkat SMFR {perangkat} Bulan {bulan_tahun_obs}, 
    terlampir kami sampaikan laporan pelaksanaan kegiatan dimaksud untuk 
    SMFR {perangkat} Site {selected_kec}, {selected_kab}.<br/><br/>
    
    Demikian disampaikan, mohon arahan lebih lanjut dan atas perhatian Bapak 
    diucapkan terimakasih.
    """
    content.append(Paragraph(isi, style_normal))
    content.append(Spacer(1, 80))
        
    # === TTD ===
    style_center_block = ParagraphStyle(name="CenterBlock", parent=styles["Normal"], alignment=TA_CENTER, fontName="Arialbd", fontSize=12)
    
    pelaksana_list = request.form.getlist("pelaksana")
    
    for p in pelaksana_list:
        pelaksana_table = Table(
            [["", Paragraph(p, style_center_block)]],
            colWidths=[350,150]  # sesuaikan lebar halaman
        )
        pelaksana_table.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "RIGHT")
        ]))
        content.append(pelaksana_table)
        content.append(Spacer(1, 50))
    
    # Build PDF
    doc.build(content)
    
    return send_file(filename, as_attachment=True, download_name="Nota_Dinas.pdf")

@app.route("/download_excel", methods=["POST"])
def download_excel():
    selected_year = request.form.get("year", "2025")
    df = load_data(selected_year)

    # Transformasi jenis identifikasi
    df["observasi_status_identifikasi_name"] = df["observasi_status_identifikasi_name"].str.replace(
        r"OFF AIR \(Sedang Tidak Digunakan\)", "OFF AIR", regex=True
    )
    df['jenis'] = df['observasi_status_identifikasi_name'].apply(
        lambda x: 'Belum Teridentifikasi' if x == 'BELUM DIKETAHUI' else 'Teridentifikasi'
    )

    # Ambil filter dari form
    selected_spt = request.form.get("spt", "Semua")
    selected_kab = request.form.get("kab", "Semua")
    selected_kec = request.form.get("kec", "Semua")
    selected_cat = request.form.get("cat", "Semua")

    # Filter data
    filt = df.copy()
    if selected_spt != "Semua":
        filt = filt[filt["observasi_no_spt"] == selected_spt]
    if selected_kab != "Semua":
        filt = filt[filt["observasi_kota_nama"] == selected_kab]
    if selected_kec != "Semua":
        filt = filt[filt["observasi_kecamatan_nama"] == selected_kec]
    if selected_cat != "Semua":
        filt = filt[filt["scan_catatan"] == selected_cat]

    if filt.empty:
        return "<h3>Data kosong, tidak bisa disimpan.</h3>"
    
    # ===============================
    # DATA KHUSUS ISR TAHUNAN
    # (hanya berdasarkan tahun & kab)
    # ===============================
    df_tahunan = df.copy()
    
    # Filter hanya berdasarkan Kab/Kota (jika dipilih)
    if selected_kab != "Semua":
        df_tahunan = df_tahunan[
            df_tahunan["observasi_kota_nama"].astype(str).str.strip().str.upper()
            == selected_kab.strip().upper()
        ]

    # Ringkasan
    jumlah_kota = filt['observasi_kota_nama'].nunique()
    legal = filt.groupby(['observasi_kota_nama', 'observasi_status_identifikasi_name']).size().reset_index(name='Jumlah')
    band = filt.pivot_table(index=['observasi_kota_nama', 'band_nama', 'jenis'], aggfunc='size', fill_value=0)
    dinas = filt.pivot_table(index=['observasi_kota_nama', 'observasi_service_name', 'jenis'], aggfunc='size', fill_value=0)
    pita = filt.pivot_table(index=['observasi_kota_nama', 'observasi_range_frekuensi', 'observasi_status_identifikasi_name'], aggfunc='size', fill_value=0)

    try:
        # === Load ISR & Samakan Format ===
        base_dir = os.path.dirname(os.path.abspath(__file__))
        isr_path = os.path.join(base_dir, "Data Target Monitor ISR 2025 - Mataram.csv")
    
        # Load CSV
        df_ISR = pd.read_csv(isr_path, on_bad_lines='skip', delimiter=';') \
                    .rename(columns={'Freq': 'Frekuensi', 'Clnt Name': 'Identifikasi'})
        df_ISR['Frekuensi'] = pd.to_numeric(df_ISR['Frekuensi'], errors='coerce')
        filt['observasi_frekuensi'] = pd.to_numeric(filt['observasi_frekuensi'], errors='coerce')
    
        # Filter ISR bila kab dipilih
        if selected_kab != "Semua":
            df_ISR = df_ISR[df_ISR['Kab/Kota'].astype(str).str.strip().str.upper() == selected_kab.strip().upper()]
    
        # === Hitung kesesuaian dengan ISR ===
        freq_df1 = filt.groupby(
            ['observasi_frekuensi','observasi_sims_client_name','observasi_kota_nama']
        ).size().reset_index(name='Jumlah_df1')
    
        freq_df1 = freq_df1.rename(columns={
            'observasi_frekuensi': 'Frekuensi',
            'observasi_sims_client_name': 'Identifikasi',
            'observasi_kota_nama': 'Kab/Kota'
        })

        freq_df2 = df_ISR.groupby(['Frekuensi','Kab/Kota']).size().reset_index(name='Jumlah_df2')
        
        # Hapus duplikat, hanya sisakan satu per kombinasi Frekuensi & Identifikasi
        merged = pd.merge(freq_df1, freq_df2, on=['Frekuensi','Kab/Kota'], how='inner')
        
        jumlah_sesuai_isr = len(merged)

        # Ambil kota yang termonitor dari data observasi
        kota_termonitor = filt['observasi_kota_nama'].dropna().astype(str).str.strip().str.upper().unique()

        # Filter target ISR berdasarkan kota termonitor
        target_match = df_ISR[df_ISR['Kab/Kota'].str.strip().str.upper().isin(kota_termonitor)]
    
        # Hitung target ISR total (jumlah baris, karena tidak ada kolom "Jumlah ISR")
        jumlah_target_isr = len(target_match) if not target_match.empty else 0

        # Hitung persen kesesuaian
        persen_sesuai_isr = round((jumlah_sesuai_isr / jumlah_target_isr * 100), 2) if jumlah_target_isr > 0 else 0
        if persen_sesuai_isr > 100:
            persen_sesuai_isr = 100
        else:
            persen_sesuai_isr = persen_sesuai_isr
    
    except Exception as e:
        print("Gagal hitung kesesuaian ISR:", e)
        jumlah_sesuai_isr, jumlah_target_isr, persen_sesuai_isr = 0, 0, 0

    # ===============================
    # HITUNG ISR TAHUNAN (TIDAK TERGANTUNG SPT)
    # ===============================
    try:
        # Samakan tipe data
        df_tahunan["observasi_frekuensi"] = pd.to_numeric(
            df_tahunan["observasi_frekuensi"], errors="coerce"
        )
    
        # Ambil kota termonitor dalam setahun
        kota_tahunan = (
            df_tahunan["observasi_kota_nama"]
            .dropna()
            .astype(str)
            .str.strip()
            .str.upper()
            .unique()
        )
    
        # Filter target ISR berdasarkan kota tahunan
        isr_tahunan = df_ISR[
            df_ISR["Kab/Kota"].astype(str).str.strip().str.upper().isin(kota_tahunan)
        ]
    
        # Observasi unik tahunan (hindari duplikat SPT)
        obs_tahunan = (
            df_tahunan
            .groupby(["observasi_frekuensi", "observasi_sims_client_name", "observasi_kota_nama"])
            .size()
            .reset_index(name="Jumlah")
            .rename(columns={
                "observasi_frekuensi": "Frekuensi",
                "observasi_sims_client_name": "Identifikasi",
                "observasi_kota_nama": "Kab/Kota"
            })
        )
    
        # Target ISR unik tahunan
        isr_tahunan_grp = (
            isr_tahunan
            .groupby(["Frekuensi", "Kab/Kota"])
            .size()
            .reset_index(name="Jumlah")
        )
    
        # Cocokkan
        merged_tahunan = pd.merge(
            obs_tahunan,
            isr_tahunan_grp,
            on=["Frekuensi", "Kab/Kota"],
            how="inner"
        )
    
        jumlah_sesuai_isr_tahunan = len(merged_tahunan)
        jumlah_target_isr_tahunan = len(isr_tahunan_grp)
    
        persen_sesuai_isr_tahunan = round(
            (jumlah_sesuai_isr_tahunan / jumlah_target_isr_tahunan * 100), 2
        ) if jumlah_target_isr_tahunan > 0 else 0
    
    except Exception as e:
        print("Gagal hitung ISR tahunan:", e)
        jumlah_sesuai_isr_tahunan = 0
        jumlah_target_isr_tahunan = 0
        persen_sesuai_isr_tahunan = 0


    jumlah_iden = len(filt[filt['jenis'] == 'Teridentifikasi'])
    jumlah_total = len(filt)
    persen_teridentifikasi = round((jumlah_iden / jumlah_total * 100), 2) if jumlah_total > 0 else 0

    # Buat Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Ringkasan Laporan"

    def add_centered_row(value, merge_cells=4):
        ws.merge_cells(start_row=ws.max_row + 1, start_column=1, end_row=ws.max_row + 1, end_column=merge_cells)
        cell = ws.cell(row=ws.max_row, column=1, value=value)
        cell.alignment = Alignment(horizontal='center')

    def add_labeled_row(label, value):
        ws.append([label, value])
    
    # Definisi border tipis
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Definisi border tipis
    thin_border2 = Border(
        bottom=Side(style='thin')
    )
            
    # Set orientasi dan scaling untuk siap print
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    
    # Scale agar muat halaman (70% dari ukuran normal)
    ws.page_setup.scale = 75
    
    # Atur margin
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5)

    # Mulai isi
    add_centered_row("RANGKUMAN HASIL LAPORAN SURAT TUGAS")
    add_centered_row(selected_spt)
    add_centered_row(f"{selected_kec} - {selected_kab}")
    add_centered_row("=" * 60)

    add_labeled_row("Jumlah Kab/Kota Termonitor:", jumlah_kota)
    add_centered_row("=" * 60)

    # Legalitas
    ws.append(["Rangkuman Berdasarkan Legalitas:"])
    for _, row in legal.iterrows():
        ws.append(row.tolist())
    add_centered_row("=" * 60)

    # Band
    ws.append(["Rangkuman Berdasarkan Band:"])
    band_df = band.reset_index()
    for _, row in band_df.iterrows():
        ws.append(row.tolist())
    add_centered_row("=" * 60)

    # Dinas
    ws.append(["Rangkuman Berdasarkan Dinas:"])
    dinas_df = dinas.reset_index()
    for _, row in dinas_df.iterrows():
        ws.append(row.tolist())
    add_centered_row("=" * 60)

    # Pita
    ws.append(["Rangkuman Berdasarkan Pita:"])
    pita_df = pita.reset_index()
    for _, row in pita_df.iterrows():
        ws.append(row.tolist())
    add_centered_row("=" * 60)

    # Tambahan ringkasan
    add_labeled_row("Jumlah Data Rekap:", jumlah_total)
    add_labeled_row("Jumlah Stasiun Radio Teridentifikasi:", f"{jumlah_iden} ({persen_teridentifikasi}%)")
    add_labeled_row("Jumlah Stasiun Radio Sesuai ISR:", f"{jumlah_sesuai_isr} ({persen_sesuai_isr}%)")
    add_labeled_row(
        "Jumlah Stasiun Radio Sesuai ISR (Tahunan):",
        f"{jumlah_sesuai_isr_tahunan} ({persen_sesuai_isr_tahunan}%)"
    )

    add_centered_row("=" * 60)
    
    # Atur lebar kolom otomatis
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_len + 2
            
    # === Tambahkan Data Observasi Terfilter di Sheet Baru ===
    filt_sheet2 = filt.copy()

    # Urutkan berdasarkan observasi_id
    if "observasi_id" in filt_sheet2.columns:
        filt_sheet2 = filt_sheet2.sort_values(by="observasi_frekuensi", ascending=True)

    # Tambahkan kolom "No" mulai dari 1
    filt_sheet2.insert(0, "No", range(1, len(filt_sheet2) + 1))

    # Ubah kolom Pita Frekuensi → ambil sebelum "."
    #if "observasi_range_frekuensi" in filt_sheet2.columns:
        #filt_sheet2["observasi_range_frekuensi"] = filt_sheet2["observasi_range_frekuensi"].astype(str).str.split(".").str[0]

    # Mapping nama kolom
    rename_map = {
        "no": "No",
        "observasi_tanggal": "Tanggal",
        "observasi_jam": "Waktu",
        "band_nama": "Band",
        "observasi_range_frekuensi": "Pita Frekuensi",
        "observasi_frekuensi": "Frekuensi",
        "observasi_level": "Level",
        "observasi_service_name": "Dinas",
        "observasi_subservice_name": "Subservice",
        "observasi_emisi_name": "Kelas Emisi",
        "observasi_equip_name": "Kelas Stasiun",
        "observasi_sims_client_name": "Nama Klien",
        "observasi_status_identifikasi_name": "Legalitas",
        "observasi_kelurahan_nama": "Kel/Desa",
        "observasi_kecamatan_nama": "Kecamatan",
        "observasi_kota_nama": "Kab/Kota",
        "observasi_propinsi_nama": "Provinsi",
        "observasi_scan_detail_lat": "Latitude",
        "observasi_scan_detail_long": "Longitude"
    }
    filt_sheet2 = filt_sheet2.rename(columns=rename_map)
    # Urutkan kolom sesuai urutan di rename_map
    ordered_cols = list(rename_map.values())
    filt_sheet2 = filt_sheet2[ordered_cols]

    # Hapus kolom tidak perlu
    drop_cols = [
        "observasi_jenis_perangkat","upt_nama","observasi_pita_frekuensi_name",
        "scan_catatan","observasi_no_spt","observasi_azimuth","observasi_jenis_stasiun",
        "observasi_tgl_spt","observasi_keterangan","observasi_status_request_delete",
        "observasi_radius","sims_tgl_query","sims_area_of_service","sims_station_name","jenis","observasi_id"
    ]
    filt_sheet2 = filt_sheet2.drop(columns=[c for c in drop_cols if c in filt_sheet2.columns], errors="ignore")

    # Buat sheet baru
    ws2 = wb.create_sheet(title="Data Terfilter")

    # Tulis header
    ws2.append(list(filt_sheet2.columns))
    
    # Tulis data
    for _, row in filt_sheet2.iterrows():
        ws2.append(row.tolist())

    # Atur lebar kolom otomatis
    for col in ws2.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws2.column_dimensions[col_letter].width = max_len + 2
    
    # Terapkan wrap text, rata tengah, dan border ke semua cell
    for row in ws2.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            
    # Persempit kolom tertentu menjadi setengah
    col_map = {v: k for k, v in enumerate(filt_sheet2.columns, start=1)}
    special_cols = ["Kelas Emisi", "Kelas Stasiun", "Nama Klien", "Pita Frekuensi"]
    
    for sc in special_cols:
        if sc in col_map:
            col_letter = get_column_letter(col_map[sc])
            current_width = ws2.column_dimensions[col_letter].width
            ws2.column_dimensions[col_letter].width = max(10, current_width / 2)
    
    # Set orientasi dan scaling untuk siap print
    ws2.page_setup.orientation = ws2.ORIENTATION_LANDSCAPE
    ws2.page_setup.paperSize = ws2.PAPERSIZE_A4
    
    # Scale agar muat halaman (70% dari ukuran normal)
    ws2.page_setup.scale = 50
    
    # Atur margin
    ws2.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5)
    
    # Tambahkan print titles (header row selalu tampil di atas setiap halaman)
    ws2.print_title_rows = "1:1"

    # Fill abu-abu untuk header baris pertama
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    for cell in ws2[1]:  # row pertama = header
        cell.fill = header_fill

    # Mode page break preview
    ws.sheet_view.view = "pageBreakPreview"
    ws2.sheet_view.view = "pageBreakPreview"

    # Simpan ke file sementara
    from tempfile import NamedTemporaryFile
    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.seek(0)

    # Ubah nama berdasarkan input
    safe_spt = safe_filename(selected_spt)
    safe_kab = safe_filename(selected_kab)
    safe_kec = safe_filename(selected_kec)
    filename = f"Rekap {safe_spt} {safe_kec} {safe_kab}.xlsx"

    return send_file(
        tmp.name,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/", methods=["GET", "POST"])
@login_required
def index():
    # === Load Info Inspeksi ===
    df_inspeksi, apstard_status = load_info_inspeksi()
    
    if not df_inspeksi.empty:
        total_inspeksi = int(df_inspeksi["total_inspeksi"].iloc[0].replace(",", ""))
        sudah_inspeksi = int(df_inspeksi["sudah_inspeksi"].iloc[0].replace(",", ""))
        belum_inspeksi = int(df_inspeksi["belum_inspeksi"].iloc[0].replace(",", ""))
        sesuai = int(df_inspeksi["sesuai"].iloc[0].replace(",", ""))
        tidak_sesuai = int(df_inspeksi["tidak_sesuai"].iloc[0].replace(",", ""))
        ilegal = int(df_inspeksi["ilegal"].iloc[0].replace(",", ""))
        off_air = int(df_inspeksi["off_air"].iloc[0].replace(",", ""))
    
        capaian_inspeksi = round((sudah_inspeksi / total_inspeksi * 100), 2) if total_inspeksi > 0 else 0
    
        # Data untuk chart
        pie_inspeksi_status = pd.DataFrame({
            "Status": ["Sesuai", "Tidak Sesuai", "Ilegal", "Off Air"],
            "Jumlah": [sesuai, tidak_sesuai, ilegal, off_air]
        })
    
        # Data untuk chart
        bar_inspeksi_status = pd.DataFrame({
            "Status": ["Tidak Sesuai", "Ilegal"],
            "Jumlah": [tidak_sesuai, ilegal]
        })
    
        pie_inspeksi = px.pie(
            pie_inspeksi_status, names="Status", values="Jumlah",
            title="Distribusi Hasil Inspeksi",
            hole=0.4,
            color_discrete_sequence=["#006db0", "#00ade6", "#edbc1b", "#8f181b", "#EF4444", "#6B7280",
                                     "#6d98b3", "#91cfe3", "#af8703", "#a83639", "#575759", "#252526",
                                     "#044065", "#d5ad2b", "#884a4c"]
        )
        pie_inspeksi.update_layout(legend=dict(orientation="v", x=1, y=0.5))
    
        bar_inspeksi = px.bar(
            bar_inspeksi_status, x="Status", y="Jumlah", color="Status",
            title="Jumlah Pelanggaran Inspeksi",
            color_discrete_sequence=["#edbc1b", "#8f181b", "#006db0", "#00ade6", "#EF4444", "#6B7280",
                                     "#6d98b3", "#91cfe3", "#af8703", "#a83639", "#575759", "#252526",
                                     "#044065", "#d5ad2b", "#884a4c"]
        )
        
        apstard_status = None  # Tidak ada error
    else:
        total_inspeksi = sudah_inspeksi = capaian_inspeksi = 0
        
        # Figure kosong biar tetap kompatibel
        pie_inspeksi = go.Figure()
        bar_inspeksi = go.Figure()
        
        # Opsional: kasih text "Data tidak tersedia"
        pie_inspeksi.add_annotation(
            text="Data tidak tersedia",
            showarrow=False,
            font=dict(size=16)
        )
        bar_inspeksi.add_annotation(
            text="Data tidak tersedia",
            showarrow=False,
            font=dict(size=16)
        )
    
        apstard_status = "⚠️ Aplikasi Apstard tidak bisa diakses"

    ############PENERTIBANNNNNNNNNN  
    pantib_selected_year = request.form.get("year", "2025")
    df_pantib = load_pantib(pantib_selected_year, use_cache=True)
    
    # Jika data kosong, fallback ke 2025
    if df_pantib is None or df_pantib.empty:
        fallback_year = "2025"
        df_pantib = load_pantib(fallback_year, use_cache=True)
        pantib_selected_year = fallback_year


    # Card 1: jumlah pelanggaran
    jumlah_pelanggaran = len(df_pantib)

    # Card 2: persentase telah ditertibkan
    total_data = len(df_pantib)
    sudah_ditertibkan = df_pantib["penertiban_no_teguran"].notna().sum()
    persentase_ditertibkan = round((sudah_ditertibkan / total_data) * 100, 2) if total_data > 0 else 0

    selected_year = request.form.get("year", "2026")
    df = load_data(selected_year)
    # Tambahkan kolom jenis
    if 'observasi_status_identifikasi_name' in df.columns:
        df['jenis'] = df['observasi_status_identifikasi_name'].apply(
            lambda x: 'Belum Teridentifikasi' if x == 'BELUM DIKETAHUI' else 'Teridentifikasi'
        )
    else:
        df['jenis'] = None  # atau default lain

    if df.empty:
        return "Data tidak tersedia. Periksa koneksi API atau cookie."

    # Ambil pilihan dari form
    selected_spt = request.form.get("spt", "Semua")
    selected_kab = request.form.get("kab", "Semua")
    selected_kec = request.form.get("kec", "Semua")
    selected_cat = request.form.get("cat", "Semua")

    # ===== Filter bertingkat =====
    if selected_spt != "Semua":
        df_spt = df[df["observasi_no_spt"] == selected_spt]
    else:
        df_spt = df

    if selected_kab != "Semua":
        df_kab = df_spt[df_spt["observasi_kota_nama"] == selected_kab]
    else:
        df_kab = df_spt

    if selected_kec != "Semua":
        df_kec = df_kab[df_kab["observasi_kecamatan_nama"] == selected_kec]
    else:
        df_kec = df_kab

    # ===== Dropdown options =====
    spt_options = ["Semua"] + sorted(df["observasi_no_spt"].dropna().unique().tolist())
    kab_options = ["Semua"] + sorted(df_spt["observasi_kota_nama"].dropna().unique().tolist())
    kec_options = ["Semua"] + sorted(df_kab["observasi_kecamatan_nama"].dropna().unique().tolist())
    cat_options = ["Semua"] + sorted(df_kec["scan_catatan"].dropna().unique().tolist())

    # ===== Filter akhir untuk tampilan data =====
    filt = df.copy()
    if selected_spt != "Semua":
        filt = filt[filt["observasi_no_spt"] == selected_spt]
    if selected_kab != "Semua":
        filt = filt[filt["observasi_kota_nama"] == selected_kab]
    if selected_kec != "Semua":
        filt = filt[filt["observasi_kecamatan_nama"] == selected_kec]
    if selected_cat != "Semua":
        filt = filt[filt["scan_catatan"] == selected_cat]

    if filt.empty:
        return f"<h3>Data kosong untuk kombinasi tersebut.</h3><p>SPT: {selected_spt}, Kab: {selected_kab}, Kec: {selected_kec}</p>"
    
    # === Ringkasan Data untuk Info Cards ===
    total_data = len(filt)
    
    # Hitung jumlah ilegal (TANPA IZIN + KADALUARSA)
    filt_no_netral = filt[~filt["observasi_status_identifikasi_name"].isin(["CLEAR", "NOISE", "IDENTIFIKASI LEBIH LANJUT", "BELUM DIKETAHUI"])]
    total_no_netral = len(filt_no_netral)
    ilegal_count = filt[filt['observasi_status_identifikasi_name'].str.upper().isin(['TANPA IZIN', 'KADALUARSA'])].shape[0]
    ilegal_percent = round((ilegal_count / total_no_netral) * 100, 2) if total_no_netral else 0
    
    # Persentase berizin = 100 - ilegal
    berizin_percent = 100 - ilegal_percent

    # Hitung jumlah & persentase off air
    offair_count = filt[filt['observasi_status_identifikasi_name'].str.upper().str.contains('OFF AIR')].shape[0]
    offair_percent = round((offair_count / total_data) * 100, 1) if total_data else 0
    
    # Hitung teridentifikasi
    teridentifikasi_count = filt[filt['jenis'] == 'Teridentifikasi'].shape[0]
    
    # Hitung jumlah Kab/Kota termonitor
    jumlah_kota_termonitor = df['observasi_kota_nama'].nunique()
    
    # Hitung persentase kab/kota termonitor
    persen_kota_termonitor = int((jumlah_kota_termonitor / 10) * 100)

    
    try:
        # === Load ISR & Samakan Format ===
        base_dir = os.path.dirname(os.path.abspath(__file__))
        isr_path = os.path.join(base_dir, "Data Target Monitor ISR 2025 - Mataram.csv")
    
        # Load CSV
        df_ISR = pd.read_csv(isr_path, on_bad_lines='skip', delimiter=';') \
                    .rename(columns={'Freq': 'Frekuensi', 'Clnt Name': 'Identifikasi'})
        df_ISR['Frekuensi'] = pd.to_numeric(df_ISR['Frekuensi'], errors='coerce')
        filt['observasi_frekuensi'] = pd.to_numeric(filt['observasi_frekuensi'], errors='coerce')
    
        # Filter ISR bila kab dipilih
        if selected_kab != "Semua":
            df_ISR = df_ISR[df_ISR['Kab/Kota'].astype(str).str.strip().str.upper() == selected_kab.strip().upper()]
    
        # === Hitung kesesuaian dengan ISR ===
        freq_df1 = filt.groupby(
            ['observasi_frekuensi','observasi_sims_client_name','observasi_kota_nama']
        ).size().reset_index(name='Jumlah_df1')
    
        freq_df1 = freq_df1.rename(columns={
            'observasi_frekuensi': 'Frekuensi',
            'observasi_sims_client_name': 'Identifikasi',
            'observasi_kota_nama': 'Kab/Kota'
        })

        freq_df2 = df_ISR.groupby(['Frekuensi','Kab/Kota']).size().reset_index(name='Jumlah_df2')
        
        # Hapus duplikat, hanya sisakan satu per kombinasi Frekuensi & Identifikasi
        merged = pd.merge(freq_df1, freq_df2, on=['Frekuensi','Kab/Kota'], how='inner')
        
        jumlah_sesuai_isr = len(merged)

        # Ambil kota yang termonitor dari data observasi
        kota_termonitor = filt['observasi_kota_nama'].dropna().astype(str).str.strip().str.upper().unique()

        # Filter target ISR berdasarkan kota termonitor
        target_match = df_ISR[df_ISR['Kab/Kota'].str.strip().str.upper().isin(kota_termonitor)]
    
        # Hitung target ISR total (jumlah baris, karena tidak ada kolom "Jumlah ISR")
        jumlah_target_isr = len(target_match) if not target_match.empty else 0
        
        # Hitung persen kesesuaian
        persen_sesuai_isr = round((jumlah_sesuai_isr / jumlah_target_isr * 100), 2) if jumlah_target_isr > 0 else 0
        if persen_sesuai_isr > 100:
            persen_sesuai_isr = 100
        else:
            persen_sesuai_isr = persen_sesuai_isr
    
    except Exception as e:
        print("Gagal hitung kesesuaian ISR:", e)
        jumlah_sesuai_isr, jumlah_target_isr, persen_sesuai_isr = 0, 0, 0


    # Persentase ISR sesuai target (sementara contoh statis)
    isr_percent = persen_sesuai_isr

    # Chart
    pie1 = px.pie(filt, names="observasi_status_identifikasi_name", title="Distribusi Legalitas",
                  hole=0.5, 
                  color_discrete_sequence=["#006db0", "#00ade6", "#edbc1b", "#8f181b", "#EF4444", "#6B7280",
                                           "#6d98b3", "#91cfe3", "#af8703", "#a83639", "#575759", "#252526",
                                           "#044065", "#d5ad2b", "#884a4c"])
    pie1.update_layout(
        legend=dict(
            font=dict(color='white', size=8))
        )
    pie_band = px.pie(filt, names="band_nama", title="Distribusi Band", 
                      hole=0.5,
                      color_discrete_sequence=["#006db0", "#00ade6", "#edbc1b", "#8f181b", "#EF4444", "#6B7280",
                                               "#6d98b3", "#91cfe3", "#af8703", "#a83639", "#575759", "#252526",
                                               "#044065", "#d5ad2b", "#884a4c"])

    bar = filt.groupby(["observasi_service_name", "jenis"]).size().reset_index(name="jumlah").sort_values(by="jumlah", ascending=False)
    total_per_dinas = (
    filt.groupby("observasi_service_name")
    .size()
    .sort_values(ascending=False))
    
    ordered_dinas = total_per_dinas.index.tolist()  # urutan berdasarkan total
    bar1 = px.bar(bar, y="observasi_service_name", x="jumlah", color="jenis", orientation='h',
                  title="Distribusi Dinas & Jenis",
                  labels={"observasi_service_name": "Nama Dinas",
                          "jumlah": "Jumlah Data",
                          "jenis": ""},
                  category_orders={"observasi_service_name": ordered_dinas},
                  color_discrete_sequence=["#006db0", "#00ade6", "#edbc1b", "#8f181b", "#EF4444", "#6B7280",
                                           "#6d98b3", "#91cfe3", "#af8703", "#a83639", "#575759", "#252526",
                                           "#044065", "#d5ad2b", "#884a4c"])
    
    bar1.update_layout(
        uniformtext_minsize=8,
        uniformtext_mode='hide',
        bargap=0.2,
        plot_bgcolor='#0f172a',
        paper_bgcolor='#0f172a',
        font=dict(color='white'),
        legend=dict(
            orientation="h",      # horizontal
            yanchor="bottom",     # anchor ke bawah
            y=-0.3,               # posisi agak di luar bawah chart
            xanchor="center",     
            x=0.5                 # posisi di tengah
        )
    )

    bar_pita = filt.groupby(["observasi_range_frekuensi", "observasi_status_identifikasi_name"]).size().reset_index(name="jumlah").sort_values(by="jumlah", ascending=False)
    # Buat kolom pita_singkat
    bar_pita["pita_singkat"] = bar_pita["observasi_range_frekuensi"].astype(str).str.split('.').str[0]

    total_per_pita = (
    filt.groupby("observasi_range_frekuensi")
    .size()
    .sort_values(ascending=False))
    
    ordered_pita = total_per_pita.index.tolist()  # urutan berdasarkan total
    # Urutan kategori singkat berdasarkan urutan original
    ordered_pita_singkat = [
        str(val).split('.')[0] for val in ordered_pita
    ]
    
    bar1_pita = px.bar(
        bar_pita,
        y="pita_singkat",
        x="jumlah",
        color="observasi_status_identifikasi_name",
        orientation='h',
        title="Distribusi Pita & Legalitas",
        labels={
            "pita_singkat": "Pita Frekuensi",
            "jumlah": "Jumlah Data",
            "observasi_status_identifikasi_name": ""
        },
        category_orders={"pita_singkat": ordered_pita_singkat},
        hover_name="observasi_range_frekuensi",
        color_discrete_sequence=["#006db0", "#00ade6", "#edbc1b", "#8f181b", "#EF4444", "#6B7280",
                                 "#6d98b3", "#91cfe3", "#af8703", "#a83639", "#575759", "#252526",
                                 "#044065", "#d5ad2b", "#884a4c"]
    )
    
    bar1_pita.update_traces(text=None)    
    bar1_pita.update_layout(
        uniformtext_minsize=8,
        uniformtext_mode='hide',
        bargap=0.2,
        plot_bgcolor='#0f172a',
        paper_bgcolor='#0f172a',
        legend=dict(
            font=dict(size=8),
            orientation="h",
            yanchor="bottom",
            y=-0.4,
            xanchor="center",
            x=0.5,
        )
    )
    
    pie_pantib = px.pie(
        df_pantib, 
        names="status_pelanggaran_name", 
        title="Jenis Pelanggaran",
        hole=0.5, 
        color_discrete_sequence=["#006db0", "#00ade6", "#edbc1b", "#8f181b", "#EF4444", "#6B7280",
                                 "#6d98b3", "#91cfe3", "#af8703", "#a83639", "#575759", "#252526",
                                 "#044065", "#d5ad2b", "#884a4c"])
    
    pie_pantib.update_layout(
        legend=dict(
            orientation="v",
            yanchor="middle",
            y=0.5,
            xanchor="left",
            x=1.05,
            font=dict(size=12)
        )
    )


    bar_data = df_pantib.groupby("status_pelanggaran_name").size().reset_index(name="jumlah")

    bar_pantib = px.bar(
        bar_data,
        x="status_pelanggaran_name",
        y="jumlah",
        title="Jumlah Pelanggaran",
        text="jumlah",
        labels={"status_pelanggaran_name":"Jenis Pelanggaran"},
        color_discrete_sequence=["#006db0", "#00ade6", "#edbc1b", "#8f181b", "#EF4444", "#6B7280",
                                 "#6d98b3", "#91cfe3", "#af8703", "#a83639", "#575759", "#252526",
                                 "#044065", "#d5ad2b", "#884a4c"]
    )
    """"
    # ===== INVOICE DENDA =====
    urllib3.disable_warnings()

    url = "https://dendaadministratif.postel.go.id/application/invoice/get/data/management"

    cookies = {
        "XSRF-TOKEN": "eyJpdiI6InNYY3UwV1VhS1RsTXFCb2FOd0cxeFE9PSIsInZhbHVlIjoiNFh4WTE5VVBZWWFWL0tKbUxYUi95R2FHWTRLd0U4THgvQnY4eGt1SjRHbkZFQ0hKcTNZTk9RcUFiMEFrOHR1T3hqOWhkTFNjeXEzdGVpZEJ4RjEzcUdYSE42dXk0d0o0UzA1QlFHTEZSbEs1bHJxT3RBY3NjOVl6aUp4b3ZaVWwiLCJtYWMiOiI2MTcxZjQzMmVhZDdlNmE1Y2ZkYWMzMzQzYjg0YzhmNjI0YTg3YTEwYmNkMzc4NWU0YTczZDdkZDBkMGY3MGUzIiwidGFnIjoiIn0%3D%3D",
        "bbppt_session": "eyJpdiI6IlozL1FsWVgremIwTzhwbzEra1Q0bUE9PSIsInZhbHVlIjoiTVNZVlFMUGdLYWltM1pQR0gxZ0pjZWJ5SkxITXJuWGoyNVc1STdRMGNGY0RnakFHZDY4cjcvV2s4d0RDWThUL3doZWVIaU5iejhIN3MyRkNjL1ZlR25wdGZCemVaZGV0T3NIUytWaHVIRjREaTAxMHgyRHZCVi9yZFdKNXk0eW0iLCJtYWMiOiIzY2I5ZWYxNzcyYTc0ZDJlMTA4ZjQ1ODFlMDZmMzJjN2E0MTdmMTA5ZGQ2NjdiYjY1NzBmZjVmMTNkNDkzMzFiIiwidGFnIjoiIn0%3D%3D"
    }

    xsrf_token = urllib.parse.unquote(cookies["XSRF-TOKEN"])

    headers = {
        "User-Agent": "Mozilla/5.0",
        "Accept": "application/json",
        "X-Requested-With": "XMLHttpRequest",
        "X-XSRF-TOKEN": xsrf_token,
        "Referer": "https://dendaadministratif.postel.go.id/application/invoice/management",
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8"
    }
    
    payload = {
        "draw": 1,
        "start": 0,
        "length": 1000,

        "columns[0][data]": "invoice_id",
        "columns[0][name]": "invoice_id",
        "columns[0][searchable]": "true",
        "columns[0][orderable]": "true",
        "columns[0][search][value]": "",
        "columns[0][search][regex]": "false",

        "columns[1][data]": "invoice_number",
        "columns[1][name]": "invoice_number",
        "columns[1][searchable]": "true",
        "columns[1][orderable]": "true",
        "columns[1][search][value]": "",
        "columns[1][search][regex]": "false",

        "order[0][column]": 0,
        "order[0][dir]": "desc",

        "search[value]": "",
        "search[regex]": "false"
    }

    resp = requests.post(
        url,
        headers=headers,
        cookies=cookies,
        data=payload,
        verify=False,
        timeout=30
    )

    print(resp.status_code)
    print(resp.text[:300])

    resp.raise_for_status()

    df_denda = pd.DataFrame(resp.json()["data"])
    df_denda.to_csv("invoice.csv", index=False)
    """
    
    df_invoice = pd.read_csv("invoice.csv")
    
    # default output
    jumlah_invoice = 0
    denda_terbayar = "Rp.0"
    denda_belum = "Rp.0"
    
    if not df_invoice.empty:
    
        # 1️⃣ Hapus baris dengan invoice_status = "Expired"
        df_invoice = df_invoice[df_invoice["invoice_status"] != "Expired"]
    
        # 2️⃣ Filter berdasarkan tahun publish_date
        # ambil 4 digit tahun dari belakang
        df_invoice["tahun"] = df_invoice["publish_date"].astype(str).str[-4:]
        
        selected_year = request.form.get("year", "2026")
        df_invoice = df_invoice[df_invoice["tahun"] == str(selected_year)]
    
        jumlah_invoice = len(df_invoice)
    
        # 3️⃣ Konversi total_amount ke integer
        df_invoice["total_amount_int"] = df_invoice["total_amount"].apply(rupiah_to_int)
    
        # 4️⃣ Hitung denda terbayar & belum
        total_terbayar = df_invoice.loc[
            df_invoice["invoice_status"] == "PAID",
            "total_amount_int"
        ].sum()
    
        total_belum = df_invoice.loc[
            df_invoice["invoice_status"] == "UNPAID",
            "total_amount_int"
        ].sum()
    
        # 5️⃣ Konversi kembali ke Rupiah
        denda_terbayar = int_to_rupiah(total_terbayar)
        print(denda_terbayar)
        denda_belum = int_to_rupiah(total_belum)
        print(denda_belum)

    
    # Data untuk Pie Chart (status pembayaran)
    pie_denda_status = pd.DataFrame({
        "Status": ["Denda Terbayar", "Belum Terbayar"],
        "Jumlah": [total_terbayar, total_belum]
    })
    pie_denda = px.pie(
        pie_denda_status,
        names="Status",
        values="Jumlah",
        title="Status Pembayaran Denda",
        hole=0.5,
        color_discrete_sequence=["#00ade6", "#edbc1b"]
    )
    pie_denda.update_layout(
        paper_bgcolor="#1e293b",
        plot_bgcolor="#1e293b",
        font=dict(color="white"),
        legend=dict(orientation="h", yanchor="top", y=-0.2, xanchor="center", x=0.5)
    )
    
    # Data untuk Bar Chart (berdasarkan dinas)
    bar_denda_dinas = pd.DataFrame({
        "Dinas": ["Point to Point", "Point to Multipoint", "Bergerak Darat"],
        "Jumlah": [3, 6, 21]
    })
    bar_denda = px.bar(
        bar_denda_dinas,
        x="Dinas",
        y="Jumlah",
        title="Denda Berdasarkan Dinas",
        text="Jumlah",
        color_discrete_sequence=["#00ade6"]
    )
    bar_denda.update_layout(
        paper_bgcolor="#1e293b",
        plot_bgcolor="#1e293b",
        font=dict(color="white"),
        margin=dict(l=40, r=20, t=60, b=80),
        legend=dict(orientation="h", yanchor="top", y=-0.2, xanchor="center", x=0.5)
    )
    
    # Convert ke JSON untuk ditampilkan di template
    pie_denda_json = json.dumps(pie_denda, cls=plotly.utils.PlotlyJSONEncoder)
    bar_denda_json = json.dumps(bar_denda, cls=plotly.utils.PlotlyJSONEncoder)

    for fig in [pie1]:
        for fig in [pie1, pie_band, bar1, bar1_pita, pie_pantib, bar_pantib, pie_inspeksi, bar_inspeksi, pie_denda, bar_denda]:
            fig.update_layout(
                paper_bgcolor="#1e293b",  # background luar chart
                plot_bgcolor="#1e293b",   # background area plot
                font=dict(color="white")  # teks jadi putih
            )
        
    for fig in [pie_band, bar1, bar1_pita, pie_pantib, bar_pantib, pie_inspeksi, bar_inspeksi, pie_denda, bar_denda]:
        fig.update_layout(
            paper_bgcolor="#1e293b",
            plot_bgcolor="#1e293b",
            font=dict(color="white"),
            margin=dict(l=40, r=20, t=60, b=80),
            legend=dict(
                orientation="h",     # horizontal
                yanchor="top",       # anchor ke atas legend
                y=-0.2,              # posisikan sedikit di bawah chart
                xanchor="center",
                x=0.5
            )
        )

    # URL Google Sheet (ubah ke format export CSV)
    sheet_id_qos = "14rFJPrA2fCVkz-7mQoLQ8khV5nLlsKVv"
    gid_qos = "301785538"
    
    url_qos = f"https://docs.google.com/spreadsheets/d/{sheet_id_qos}/export?format=csv&gid={gid_qos}"
    
    # Ambil data ke DataFrame
    df_qos = pd.read_csv(url_qos)
    df_qos = pd.read_csv(url_qos, header=1)
    jumlah_qos = df_qos['Kabupaten / Kota'].nunique()
    persen_qos = int(jumlah_qos/10*100)

    df_map = load_map_data()
    map_static_file = generate_map_html_from_df(df_map)

    # =========================
    # DATA UNTUK CARD SPEEDTEST
    # =========================
    
    speed_cols = [
        "Average Speedtest Download Speed (Mbps)",
        "Average Speedtest Upload Speed (Mbps)"
    ]
    
    # Paksa konversi ke numeric (non-angka → NaN)
    for col in speed_cols:
        df_map[col] = (
            df_map[col]
            .astype(str)
            .str.replace(",", ".", regex=False)
            .str.strip()
        )
        df_map[col] = pd.to_numeric(df_map[col], errors="coerce")
    
    # Hapus baris yang tidak punya nilai speed valid
    operator_speed_df = df_map.dropna(
        subset=[
            "Operator",
            "Average Speedtest Download Speed (Mbps)",
            "Average Speedtest Upload Speed (Mbps)"
        ]
    )
    
    # Ambil kolom yang dibutuhkan saja
    operator_speed_df = operator_speed_df[[
        "Operator",
        "Average Speedtest Download Speed (Mbps)",
        "Average Speedtest Upload Speed (Mbps)"
    ]]
    
    # Convert ke JSON
    operator_speed_json = operator_speed_df.to_dict(orient="records")

    
    pie1_json = json.dumps(pie1, cls=plotly.utils.PlotlyJSONEncoder)
    pie_band_json = json.dumps(pie_band, cls=plotly.utils.PlotlyJSONEncoder)
    bar1_json = json.dumps(bar1, cls=plotly.utils.PlotlyJSONEncoder)
    bar1_pita_json = json.dumps(bar1_pita, cls=plotly.utils.PlotlyJSONEncoder)
    pie_pantib_json = json.dumps(pie_pantib, cls=PlotlyJSONEncoder)
    bar_pantib_json = json.dumps(bar_pantib, cls=PlotlyJSONEncoder)
    
    pie_inspeksi_json = json.dumps(pie_inspeksi, cls=PlotlyJSONEncoder)
    bar_inspeksi_json = json.dumps(bar_inspeksi, cls=PlotlyJSONEncoder)
    
    return render_template_string('''
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>REALISASI BALAI MONITOR SFR KELAS II MATARAM</title>
        <link rel="icon" type="image/png" href="{{ url_for('static', filename='D.png') }}">
        <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
        <style>
            body { font-family: Arial; padding: 20px; }
            select { margin-right: 10px; }
            .chart { margin-top: 30px; }
            .chart-row {
                display: flex;
                flex-wrap: wrap;     /* biar chart bisa turun ke bawah kalau layar sempit */
                gap: 20px;
                margin: 20px 0;
            }
            }
            .chart {
                flex: 1;
                width: 100%;       /* ikut lebar container */
                height: auto;      /* tinggi menyesuaikan */
                min-width: 0;      /* biar bisa mengecil */
            }
            body {
                background-color: #0d1b2a;
                color: white;
                font-family: 'Segoe UI', sans-serif;
            }
        
            .filter-form {
                display: flex;
                flex-wrap: wrap;
                gap: 20px;
                margin: 2px 0;
                background-color: #1e1e1e; /* gelap */
                padding: 20px;
                border-radius: 10px;
                box-shadow: 0 2px 6px rgba(0, 0, 0, 0.3);
            }
            
            .filter-form2 {
                display: flex;
                flex-wrap: wrap;
                gap: 1px;
                margin: 1px 0;
                padding: 0px;
                border-radius: 1px;
                box-shadow: 0 2px 6px rgba(0, 0, 0, 0.3);
            }
            
            .filter-form3 {
                display: flex;
                flex-wrap: wrap;
                gap: 10px;
                margin: 2px 0;
                background-color: #1e1e1e; /* gelap */
                padding: 10px;
                border-radius: 10px;
                box-shadow: 0 2px 6px rgba(0, 0, 0, 0.3);
                color: #e1ae05;
            }
            
            .filter-group label {
                font-weight: 600;
                margin-bottom: 5px;
                color: #ddd; /* teks terang */
            }
            
            .filter-group select {
                padding: 8px;
                border: 1px solid #444;
                border-radius: 6px;
                font-size: 14px;
                background-color: #2a2a2a; /* gelap */
                color: #f5f5f5; /* teks terang */
            }
            
            .filter-buttons button {
                padding: 10px 16px;
                background-color: #007bff;
                color: white;
                border: none;
                border-radius: 6px;
                font-weight: bold;
                cursor: pointer;
                transition: 0.3s;
            }
            
            .filter-buttons button:hover {
                background-color: #0056b3;
            }
            .chart-container {
                flex: 1 1 400px;     /* minimal 400px, tapi bisa melebar penuh */
                background-color: #1e293b;
                border-radius: 10px;
                padding: 15px;
                min-height: 400px;   /* tinggi minimal */
                width: 100%;         /* agar ikuti parent */
                position: relative;
                z-index: 1;
            }
            
            .modal {
              z-index: 9999 !important;   /* pastikan paling tinggi */
            }
            
            .modal-backdrop {
              z-index: 9998 !important;   /* backdrop di bawah modal tapi tetap di atas chart */
            }
            
            .map-container {
                width: 100%;
                height: 650px;
                margin-top: 20px;
                border-radius: 12px;
                overflow: hidden;
                box-shadow: 0 4px 20px rgba(0,0,0,0.15);
                background-color: white;
            }
            h2 {
                margin-top: 20px;
            }
        </style>
        
    </head>
    
    <script>
    function syncExcelForm() {
      document.getElementById('excel-spt').value = document.getElementById('spt').value;
      document.getElementById('excel-kab').value = document.getElementById('kab').value;
      document.getElementById('excel-kec').value = document.getElementById('kec').value;
      document.getElementById('excel-cat').value = document.getElementById('cat').value;
    }
    
    function autoSubmit(level) {
      const form = document.getElementById('main-form');
      const spt = document.getElementById('spt');
      const kab = document.getElementById('kab');
      const kec = document.getElementById('kec');
      const cat = document.getElementById('cat');
    
      // Reset bertingkat & kunci dropdown berikutnya
      if (level === 'spt') {
        kab.selectedIndex = 0; kab.disabled = false;
        kec.selectedIndex = 0; kec.disabled = true;
        cat.selectedIndex = 0; cat.disabled = true;
      } else if (level === 'kab') {
        kec.selectedIndex = 0; kec.disabled = false;
        cat.selectedIndex = 0; cat.disabled = true;
      } else if (level === 'kec') {
        cat.selectedIndex = 0; cat.disabled = false;
      }
    
      // Sinkron nilai ke form Excel
      syncExcelForm();
    
      // Auto-submit untuk refresh data di server
      // requestSubmit() akan hormati type & constraints, fallback ke submit()
      if (form.requestSubmit) form.requestSubmit();
      else form.submit();
    }
    
    // Saat halaman pertama kali dimuat, pastikan form Excel tersinkron
    document.addEventListener('DOMContentLoaded', syncExcelForm);
    </script>



    <body style="background-color:#0d1b2a; color:white; font-family:Segoe UI, sans-serif; zoom:80%;">

    <div style="display:flex; align-items:center; gap:15px; padding:20px;">
        <img src="/static/logo-komdigi2.png" style="height:50px;">
        <img src="/static/djid.png" style="height:50px;">
        <div>
            <h2 style="margin:0;">Dashboard Realisasi Kinerja – Balmon SFR Kelas II Mataram</h2/>
            <p style="margin:0; font-size:16px; color:#9ca3af;">One-Data Aggregation & Analytics (WANDAA)</p>
        </div>
        
        <!-- Tombol + Pilih Tahun -->
        <div style="display:flex; align-items:flex-end; gap:10px; padding:20px; justify-content:flex-end;">      
            <!-- Tombol -->
            <form method="POST" class="filter-form2" id="main-form">
                <!-- Pilih Tahun -->
                <div style="display:flex; flex-direction:column;">
                    <select name="year" id="year" onchange="autoSubmit('year')";
                            style="padding:10px 20px; font-size: 16px; border-radius:6px; border:none; background:#006db0; color:white;">
                        <option value="2022" {% if selected_year == "2022" %}selected{% endif %}>TAHUN 2022</option>
                        <option value="2023" {% if selected_year == "2023" %}selected{% endif %}>TAHUN 2023</option>
                        <option value="2024" {% if selected_year == "2024" %}selected{% endif %}>TAHUN 2024</option>
                        <option value="2025" {% if selected_year == "2025" %}selected{% endif %}>TAHUN 2025</option>
                        <option value="2026" {% if selected_year == "2026" %}selected{% endif %}>TAHUN 2026</option>
                    </select>
                </div>
            </form>
        
            <button form="excel-form" type="submit"
                    style="background:#006db0; color:white; border:none; padding:10px 20px; font-size: 16px; border-radius:6px;">
                Unduh Rekap
            </button>
        
            <button type="button" onclick="openModal()"
                    style="background:#e1ae05; color:white; border:none; padding:10px 20px; font-size: 16px; border-radius:6px;">
                Unduh Nodin
            </button>
        
            <a href="{{ url_for('identifikasi') }}"
                   style="color:white; background:#e1ae05; border:none; padding:10px 20px; font-size:16px; border-radius:6px; text-decoration:none;">
                    Identifikasi
                </a>
            </a>
            
            <a href="{{ url_for('run_bts') }}"
                   style="color:white; background:#e1ae05; border:none; padding:10px 20px; font-size:16px; border-radius:6px; text-decoration:none;">
                    Prima Aksi
                </a>
            </a>

            
            <a href="{{ url_for('logout') }}"
               style="color:white; background:red; border:none; padding:10px 20px; font-size: 16px; border-radius:6px;">
                Logout
            </a>
        </div>
    </div>
    
    <!-- Filter OBSMON-->
    <form method="POST" class="filter-form3" id="main-form">
    <div style="display:flex; left-content:space-between; align-items:center; margin:10px 10px 10px;">
        <img src="/static/vector.png" alt="Data" style="width:60px; height:40px;">
        <h2 style="margin:0;">MONITORING KAB/KOTA {{selected_year}}</h2>
    </div>
    </form>
    
    <form method="POST" class="filter-form" id="filter-form">
        
        <!-- Tahun -->
        <div class="filter-group">
        <select name="year" onchange="this.form.submit()">
            <option value="2022" {% if selected_year == "2022" %}selected{% endif %}>TAHUN 2022</option>
            <option value="2023" {% if selected_year == "2023" %}selected{% endif %}>TAHUN 2023</option>
            <option value="2024" {% if selected_year == "2024" %}selected{% endif %}>TAHUN 2024</option>
            <option value="2025" {% if selected_year == "2025" %}selected{% endif %}>TAHUN 2025</option>
            <option value="2026" {% if selected_year == "2026" %}selected{% endif %}>TAHUN 2026</option>
        </select>
        </div>
    
        <!-- SPT -->
        <div class="filter-group">
        <label for="spt">No SPT</label>
        <select name="spt" onchange="this.form.submit()">
            {% for spt in spt_options %}
            <option value="{{ spt }}" {% if spt == selected_spt %}selected{% endif %}>{{ spt }}</option>
            {% endfor %}
        </select>
        </div>
    
        <!-- Kab/Kota -->
        <div class="filter-group">
        <label for="kab">Kab/Kota</label>
        <select name="kab" onchange="this.form.submit()">
            {% for kab in kab_options %}
            <option value="{{ kab }}" {% if kab == selected_kab %}selected{% endif %}>{{ kab }}</option>
            {% endfor %}
        </select>
        </div>
    
        <!-- Kecamatan -->
        <div class="filter-group">
        <label for="kec">Kecamatan</label>
        <select name="kec" onchange="this.form.submit()">
            {% for kec in kec_options %}
            <option value="{{ kec }}" {% if kec == selected_kec %}selected{% endif %}>{{ kec }}</option>
            {% endfor %}
        </select>
        </div>
    
        <!-- Catatan -->
        <div class="filter-group">
        <label for="cat">Catatan</label>
        <select name="cat" onchange="this.form.submit()">
            {% for cat in cat_options %}
            <option value="{{ cat }}" {% if cat == selected_cat %}selected{% endif %}>{{ cat }}</option>
            {% endfor %}
        </select>
        </div>
    
    </form>


    <!-- Info Cards -->
    <div style="display:grid; grid-template-columns: repeat(5, 1fr); gap:15px; padding:20px;">
        
        <!-- Card 1 -->
        <div style="background:#1e293b; padding:8px; border-radius:8px; display:flex; align-items:center; gap:12px;">
            <img src="/static/1.png" alt="Data" style="width:36px; height:36px;">
            <div>
                <h1 style="margin:0;">{{ total_data }}</h1>
                <p style="margin:0;">Total Data Observasi</p>
            </div>
        </div>
        
        <!-- Card 2 -->
        <div style="background:#1e293b; padding:8px; border-radius:8px; display:flex; align-items:center; gap:12px;">
            <img src="/static/4.png" alt="Teridentifikasi" style="width:36px; height:36px;">
            <div>
                <h1 style="margin:0;">{{ teridentifikasi_count }}</h1>
                <p style="margin:0;">Teridentifikasi</p>
            </div>
        </div>
        
        <!-- Card 3 -->
        <div style="background:#1e293b; padding:8px; border-radius:8px; display:flex; align-items:center; gap:12px;">
            <img src="/static/2.png" alt="Berizin" style="width:36px; height:36px;">
            <div>
                <h1 style="margin:0;">{{ berizin_percent }}%</h1>
                <p style="margin:0;">Berizin</p>
            </div>
        </div>
        
        <!-- Card 4 -->
        <div style="background:#1e293b; padding:8px; border-radius:8px; display:flex; align-items:center; gap:12px;">
            <img src="/static/5.png" alt="ISR" style="width:36px; height:36px;">
            <div>
                <h1 style="margin:0;">{{ isr_percent }}%</h1>
                <p style="margin:0;">ISR Sesuai Target</p>
            </div>
        </div>
        
        <!-- Card 5 -->
        <div style="background:#1e293b; padding:8px; border-radius:8px; display:flex; align-items:center; gap:12px;">
            <img src="/static/6.png" alt="Kab/Kota Termonitor" style="width:36px; height:36px;">
            <div>
                <h1 style="margin:0;">{{ persen_kota_termonitor }}%</h1>
                <p style="margin:0;">Kab/Kota Termonitor</p>
            </div>
        </div>
    </div>
    
    <!-- Modal -->
    <div id="laporanModal"
         style="display:none; position:fixed; top:0; left:0; width:100%; height:100%; background:rgba(0,0,0,0.5); z-index: 99999;">
      <div style="background:white; width:400px; margin:100px auto; padding:20px; border-radius:8px; position:relative; z-index: 100000;">
        <h3>Isi Nama Pelaksana</h3>
        <!-- Form di dalam modal -->
        <form id="laporanForm" method="POST" action="/unduh_laporan">
          <!-- Hidden filter biar tetap terkirim -->
          <input type="hidden" name="spt" value="{{ selected_spt }}">
          <input type="hidden" name="kab" value="{{ selected_kab }}">
          <input type="hidden" name="kec" value="{{ selected_kec }}">
          <input type="hidden" name="cat" value="{{ selected_cat }}">
          
          <!-- Tanggal SPT -->
          <div style="margin-bottom:10px;">
            <label for="tgl_spt" style="font-weight:bold; display:block; margin-bottom:5px; color:#333;">
              📅 Tanggal SPT
            </label>
            <input type="date" name="tgl_spt" id="tgl_spt"
                   style="width:100%; padding:6px; border:1px solid #ccc; border-radius:5px;">
          </div>
          
          <!-- Pilihan Perangkat -->
            <div style="margin-bottom:10px;">
              <label for="perangkat" style="font-weight:bold; display:block; margin-bottom:5px; color:#333;">
                ⚙️ Perangkat SMFR
              </label>
              <select name="perangkat" id="perangkat"
                      style="width:100%; padding:6px; border:1px solid #ccc; border-radius:5px;">
                <option value="Tetap/Transportable TCI">Tetap/Transportable TCI</option>
                <option value="Tetap/Transportable LS Telecom">Tetap/Transportable LS Telecom</option>
                <option value="Bergerak R&S DDF205 Unit Mobil Isuzu Elf / Hilux Hitam">
                  Bergerak R&S DDF205 Unit Mobil Isuzu Elf / Hilux Hitam
                </option>
                <option value="Bergerak R&S DDF205 Unit Mobil Hilux Silver">
                  Bergerak R&S DDF205 Unit Mobil Hilux Silver
                </option>
                <option value="Jinjing R&S DDF007">Jinjing R&S DDF007</option>
                <option value="Jinjing R&S PR100">Jinjing R&S PR100</option>
              </select>
            </div>

          <!-- Nama Pelaksana -->
          <div id="pelaksana-container">
            <input type="text" name="pelaksana" placeholder="Nama Pelaksana"
                   style="width:100%; margin-bottom:10px; padding:6px;">
          </div>
    
          <!-- Tombol tambah pelaksana -->
          <button type="button" onclick="tambahPelaksana()"
                  style="background:#006db0; color:white; border:none; padding:6px 12px; border-radius:5px;">
            ➕ Tambah Pelaksana
          </button>
    
          <!-- Tombol aksi -->
          <div style="margin-top:15px; text-align:right;">
            <button type="button" onclick="closeModal()"
                    style="background:#ccc; border:none; padding:6px 12px; border-radius:5px;">
              Batal.
            </button>
            <button type="submit"
                    style="background:#006db0; color:white; border:none; padding:8px 14px; border-radius:6px;">
              ✅ Unduh
            </button>
          </div>
        </form>
      </div>
    </div>
    
    <script>
    function openModal() {
      document.getElementById("laporanModal").style.display = "block";
    }
    function closeModal() {
      document.getElementById("laporanModal").style.display = "none";
    }
    function tambahPelaksana() {
      const container = document.getElementById("pelaksana-container");
      const input = document.createElement("input");
      input.type = "text";
      input.name = "pelaksana";
      input.placeholder = "Nama Pelaksana";
      input.style = "width:100%; margin-bottom:10px; padding:6px;";
      container.appendChild(input);
    }
    </script>


    
    <!-- Form Unduh Rekap (disinkron otomatis oleh JS) -->
        <form method="POST" action="/download_excel" id="excel-form" style="display:none;">
          <input type="hidden" name="spt" id="excel-spt" value="{{ selected_spt }}">
          <input type="hidden" name="kab" id="excel-kab" value="{{ selected_kab }}">
          <input type="hidden" name="kec" id="excel-kec" value="{{ selected_kec }}">
          <input type="hidden" name="cat" id="excel-cat" value="{{ selected_cat }}">
        </form>
    
        <form method="POST" action="/unduh_laporan">
            <input type="hidden" name="spt" value="{{ selected_spt }}">
            <input type="hidden" name="kab" value="{{ selected_kab }}">
            <input type="hidden" name="kec" value="{{ selected_kec }}">
            <input type="hidden" name="cat" value="{{ selected_cat }}">
        </form>

    
    <!-- Charts -->
    <div class="chart-row">
        <div class="chart-container" id="pie1"></div>
        <div class="chart-container" id="bar1_pita"></div>      
    </div>
    
    <div class="chart-row">
        <div class="chart-container" id="pie_band"></div>
        <div class="chart-container" id="bar1"></div>
    </div>

    <!-- Info Cards Pantib -->
    <form method="POST" class="filter-form3" id="main-form">
    <div style="display:flex; left-content:space-between; align-items:center; margin:10px 10px 10px;">
        <img src="/static/vector3.png" alt="Data" style="width:60px; height:40px;">
        <h2 style="margin:0;">TEMUAN PENERTIBAN {{ pantib_selected_year }}</h2>
    </div>
    </form>
    
    <div style="display:flex; gap:15px; padding:20px; flex-wrap:wrap;">
        <div style="flex:1; min-width:200px; background:#1e293b; padding:15px; border-radius:8px; display:flex; align-items:center; gap:10px;">
            <div style="font-size:2rem;">⚠️</div>
            <div>
                <h1 style="margin:0;">{{ jumlah_pelanggaran }}</h1>
                <p>Jumlah Pelanggaran</p>
            </div>
        </div>
        <div style="flex:1; min-width:200px; background:#1e293b; padding:15px; border-radius:8px; display:flex; align-items:center; gap:10px;">
            <div style="font-size:2rem;">✅</div>
            <div>
                <h1 style="margin:0;">{{ persentase_ditertibkan }}%</h1>
                <p>Telah Ditertibkan</p>
            </div>
        </div>
    </div>

    <div class="chart-row">
        <div class="chart-container" id="pie_pantib"></div>
        <div class="chart-container" id="bar_pantib"></div>
    </div>
        
    <!-- === Chart Penanganan Denda === -->
    <form method="POST" class="filter-form3" id="main-form">
    <div style="display:flex; left-content:space-between; align-items:center; margin:10px 10px 10px;">
        <img src="/static/vector3.png" alt="Data" style="width:60px; height:40px;">
        <h2 style="margin:0;">  PENANGANAN DENDA {{selected_year}}</h2>
    </div>
    </form>
    
    <div style="
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
        gap: 20px;
        padding: 20px;
    ">
        <!-- Card Denda Terbayar -->
        <div style="
            background: #1e293b;
            padding: 20px;
            border-radius: 12px;
            text-align: center;
            box-shadow: 0 2px 6px rgba(0,0,0,0.3);
            transition: transform 0.2s ease, box-shadow 0.2s ease;
        " 
        onmouseover="this.style.transform='scale(1.03)'; this.style.boxShadow='0 4px 12px rgba(0,0,0,0.4)';"
        onmouseout="this.style.transform='scale(1)'; this.style.boxShadow='0 2px 6px rgba(0,0,0,0.3)';">
            <h1 style="margin:0; color:#00ade6;">{{denda_terbayar}}</h1>
            <p style="margin:6px 0 0; color:#e5e7eb; font-size:18px;">Denda Terbayar</p>
        </div>
    
        <!-- Card Denda Belum Terbayar -->
        <div style="
            background: #1e293b;
            padding: 20px;
            border-radius: 12px;
            text-align: center;
            box-shadow: 0 2px 6px rgba(0,0,0,0.3);
            transition: transform 0.2s ease, box-shadow 0.2s ease;
        "
        onmouseover="this.style.transform='scale(1.03)'; this.style.boxShadow='0 4px 12px rgba(0,0,0,0.4)';"
        onmouseout="this.style.transform='scale(1)'; this.style.boxShadow='0 2px 6px rgba(0,0,0,0.3)';">
            <h1 style="margin:0; color:#edbc1b;">{{denda_belum}}</h1>
            <p style="margin:6px 0 0; color:#e5e7eb; font-size:18px;">Denda Belum Terbayar</p>
        </div>
    </div>

    
    <div class="chart-row">
        <div class="chart-container" id="pie_denda"></div>
        <div class="chart-container" id="bar_denda"></div>
    </div>

    
    <!-- Info Cards Inspeksi -->
    <form method="POST" class="filter-form3" id="main-form">
    <div style="display:flex; left-content:space-between; align-items:center; margin:10px 10px 10px;">
        <img src="/static/vector.png" alt="Data" style="width:60px; height:40px;">
        <h2 style="margin:0;">PEMERIKSAAN MICROWAVE LINK {{selected_year}}</h2>
    </div>
    </form>
    <div style="display:flex; gap:15px; padding:20px; flex-wrap:wrap;">
        <div style="flex:1; min-width:200px; background:#1e293b; padding:15px; border-radius:8px; display:flex; align-items:center; gap:10px;">
            <div style="font-size:2rem;">📊</div>
            <div>
                <h1 style="margin:0;">{{ total_inspeksi }}</h1>
                <p>Total Inspeksi</p>
            </div>
        </div>
        <div style="flex:1; min-width:200px; background:#1e293b; padding:15px; border-radius:8px; display:flex; align-items:center; gap:10px;">
            <div style="font-size:2rem;">🔍</div>
            <div>
                <h1 style="margin:0;">{{ sudah_inspeksi }}</h1>
                <p>Sudah Inspeksi</p>
            </div>
        </div>
    </div>
    
    {% if apstard_status %}
        <!-- Kalau Apstard tidak bisa diakses -->
        <div style="padding:20px; background:#ef4444; color:white; border-radius:8px; margin-top:15px;">
            ⚠️ {{ apstard_status }}
        </div>
    {% else %}
        <!-- Chart hanya muncul kalau data ada -->
        <div class="chart-row">
            <div class="chart-container" id="pie_inspeksi"></div>
            <div class="chart-container" id="bar_inspeksi"></div>
        </div>
    {% endif %}

    <!-- ================= DASHBOARD QOS & SPEED ================= -->
    <form method="POST" class="filter-form3" id="main-form">
    <div style="display:flex; left-content:space-between; align-items:center; margin:10px 10px 10px;">
        <img src="/static/vector.png" alt="Data" style="width:60px; height:40px;">
        <h2 style="margin:0;">KUALITAS LAYANAN JARINGAN SELULER {{selected_year}}</h2>
    </div>
    </form>
    
    <div style="
        display:grid;
        grid-template-columns: repeat(auto-fit, minmax(240px, 1fr));
        gap:20px;
        padding:20px;
    ">
    
        <!-- CARD JUMLAH KAB/KOTA -->
        <div style="background:#1e293b; padding:18px; border-radius:12px; display:flex; gap:14px; align-items:center;">
            <div style="font-size:2.5rem;">📊</div>
            <div>
                <h1 style="margin:0;">{{ jumlah_qos }}</h1>
                <p style="margin:0; opacity:0.8;">Kab/Kota Termonitor</p>
            </div>
        </div>
    
        <!-- CARD PERSENTASE -->
        <div style="background:#1e293b; padding:18px; border-radius:12px; display:flex; gap:14px; align-items:center;">
            <div style="font-size:2.5rem;">📈</div>
            <div>
                <h1 style="margin:0;">{{ persen_qos }}%</h1>
                <p style="margin:0; opacity:0.8;">Cakupan Monitoring</p>
            </div>
        </div>
    
        <!-- CARD DOWNLOAD -->
        <div style="background:#1e293b; padding:18px; border-radius:12px;">
            <h4 style="margin-bottom:8px;">📥 Download Speed</h4>
            <h1 id="avgDownloadText">- Mbps</h1>
    
            <div style="background:#334155; border-radius:6px; height:14px; width:100%; overflow:hidden;">
                <div id="avgDownloadBar"
                     style="height:100%; width:0%; background:linear-gradient(90deg,#ef4444,#eab308,#22c55e);">
                </div>
            </div>
            <small style="opacity:0.7;">Skala hingga 100 Mbps</small>
        </div>
    
        <!-- CARD UPLOAD -->
        <div style="background:#1e293b; padding:18px; border-radius:12px;">
            <h4 style="margin-bottom:8px;">📤 Upload Speed</h4>
            <h1 id="avgUploadText">- Mbps</h1>
    
            <div style="background:#334155; border-radius:6px; height:14px; width:100%; overflow:hidden;">
                <div id="avgUploadBar"
                     style="height:100%; width:0%; background:linear-gradient(90deg,#ef4444,#eab308,#22c55e);">
                </div>
            </div>
            <small style="opacity:0.7;">Skala hingga 100 Mbps</small>
        </div>
    
        <!-- FILTER OPERATOR -->
        <div style="background:#1e293b; padding:18px; border-radius:12px;">
            <h4 style="margin-bottom:10px;">📡 Operator Seluler</h4>
            <label><input type="checkbox" class="op-filter" value="Telkomsel" checked> Telkomsel</label><br>
            <label><input type="checkbox" class="op-filter" value="Indosat" checked> Indosat</label><br>
            <label><input type="checkbox" class="op-filter" value="XL" checked> XL</label><br>
            <label><input type="checkbox" class="op-filter" value="Smart" checked> Smart</label>
        </div>
    
    </div>

    <h2 style="margin:30px 20px 10px;">🗺️ Peta Sebaran Koordinat QoS</h2>
    
    <div class="map-container" style="height:600px; margin:20px;">
        <iframe
            src="/static/{{ map_static_file }}"
            width="100%"
            height="100%"
            style="border:none; border-radius:12px;">
        </iframe>
    </div>

    <script>
        Plotly.newPlot("pie1", {{ pie1_json|safe }}.data, {{ pie1_json|safe }}.layout, {responsive: true});
        Plotly.newPlot("bar1_pita", {{ bar1_pita_json|safe }}.data, {{ bar1_pita_json|safe }}.layout, {responsive: true});
        Plotly.newPlot("pie_band", {{ pie_band_json|safe }}.data, {{ pie_band_json|safe }}.layout, {responsive: true});
        Plotly.newPlot("bar1", {{ bar1_json|safe }}.data, {{ bar1_json|safe }}.layout, {responsive: true});
    </script>
    
    <script>
        var piePantib = {{ pie_pantib_json|safe }};
        var barPantib = {{ bar_pantib_json|safe }};
        Plotly.newPlot('pie_pantib', piePantib.data, piePantib.layout, {responsive:true});
        Plotly.newPlot('bar_pantib', barPantib.data, barPantib.layout, {responsive:true});
    </script>
    
    <script>
        var pieInspeksi = {{ pie_inspeksi_json|safe }};
        var barInspeksi = {{ bar_inspeksi_json|safe }};
        Plotly.newPlot('pie_inspeksi', pieInspeksi.data, pieInspeksi.layout, {responsive:true});
        Plotly.newPlot('bar_inspeksi', barInspeksi.data, barInspeksi.layout, {responsive:true});
    </script>
    
    <script>
        var pieInspeksi = {{ pie_denda_json|safe }};
        var barInspeksi = {{ bar_denda_json|safe }};
        Plotly.newPlot('pie_denda', {{ pie_denda_json | safe }});
        Plotly.newPlot('bar_denda', {{ bar_denda_json | safe }});
    </script>
    
    <script>
    const speedData = {{ operator_speed_json | safe }};
    
    // batas skala (bisa diubah)
    const MAX_DOWNLOAD = 100; // Mbps
    const MAX_UPLOAD = 100;   // Mbps
    
    function updateSpeedCards() {
        const selectedOperators = Array.from(
            document.querySelectorAll('.op-filter:checked')
        ).map(cb => cb.value);
    
        const filtered = speedData.filter(d =>
            selectedOperators.includes(d.Operator)
        );
    
        if (filtered.length === 0) {
            document.getElementById("avgDownloadText").innerText = "- Mbps";
            document.getElementById("avgUploadText").innerText = "- Mbps";
            document.getElementById("avgDownloadBar").style.width = "0%";
            document.getElementById("avgUploadBar").style.width = "0%";
            return;
        }
    
        const avgDL = filtered.reduce(
            (s, d) => s + d["Average Speedtest Download Speed (Mbps)"], 0
        ) / filtered.length;
    
        const avgUL = filtered.reduce(
            (s, d) => s + d["Average Speedtest Upload Speed (Mbps)"], 0
        ) / filtered.length;
    
        // Text
        document.getElementById("avgDownloadText").innerText = avgDL.toFixed(2) + " Mbps";
        document.getElementById("avgUploadText").innerText = avgUL.toFixed(2) + " Mbps";
    
        // Bar scale (%)
        const dlPercent = Math.min((avgDL / MAX_DOWNLOAD) * 100, 100);
        const ulPercent = Math.min((avgUL / MAX_UPLOAD) * 100, 100);
    
        document.getElementById("avgDownloadBar").style.width = dlPercent + "%";
        document.getElementById("avgUploadBar").style.width = ulPercent + "%";
    }
    
    // event listener
    document.querySelectorAll('.op-filter').forEach(cb =>
        cb.addEventListener('change', updateSpeedCards)
    );
    
    // init
    updateSpeedCards();
    </script>

    </body>
    
    <!-- Loading Overlay -->
    <div id="loading-overlay"
         style="
            display:none;
            position:fixed;
            top:0; left:0;
            width:100%; height:100%;
            background:rgba(0,0,0,0.6);
            z-index:9999;
            align-items:center;
            justify-content:center;
            color:white;
            font-size:18px;
         ">
    
        <div style="background:#1e1e1e; padding:30px 40px; border-radius:10px; text-align:center;">
    
            <!-- SPINNER DI SINI -->
            <div style="
                border:4px solid #444;
                border-top:4px solid #00aaff;
                border-radius:50%;
                width:40px;
                height:40px;
                animation:spin 1s linear infinite;
                margin:0 auto 15px;
            "></div>
    
            <strong>Sedang proses</strong><br>
            <span style="font-size:14px;">Mohon tunggu sebentar...</span>
        </div>
    </div>
    
    <style>
    @keyframes spin {
        0%   { transform: rotate(0deg); }
        100% { transform: rotate(360deg); }
    }
    </style>

    
    <script>
    function autoSubmit() {
        const overlay = document.getElementById("loading-overlay");
        overlay.style.display = "flex";
        document.getElementById("main-form").submit();
    }
    </script>

    </html>
    ''',
    spt_options=spt_options,
    kab_options=kab_options,
    kec_options=kec_options,
    cat_options=cat_options,
    selected_year=selected_year,
    pantib_selected_year=pantib_selected_year,
    selected_spt=selected_spt,
    selected_kab=selected_kab,
    selected_kec=selected_kec,
    selected_cat=selected_cat,
    pie1_json=pie1_json,
    pie_band_json=pie_band_json,
    bar1_json=bar1_json,
    bar1_pita_json=bar1_pita_json,
    total_data=total_data,
    berizin_percent=berizin_percent,
    offair_percent=offair_percent,
    persen_kota_termonitor = persen_kota_termonitor,
    teridentifikasi_count=teridentifikasi_count,
    isr_percent=isr_percent,
    jumlah_pelanggaran=jumlah_pelanggaran,
    persentase_ditertibkan=persentase_ditertibkan,
    pie_pantib=pie_pantib.to_html(full_html=False),
    bar_pantib=bar_pantib.to_html(full_html=False),
    pie_pantib_json=pie_pantib_json,
    bar_pantib_json=bar_pantib_json,
    sudah_inspeksi = sudah_inspeksi,
    total_inspeksi = total_inspeksi,
    capaian_inspeksi = capaian_inspeksi,
    pie_inspeksi_json=pie_inspeksi_json,
    bar_inspeksi_json=bar_inspeksi_json,
    denda_terbayar=denda_terbayar,
    denda_belum=denda_belum,
    pie_denda_json=pie_denda_json,
    bar_denda_json=bar_denda_json,
    jumlah_qos=jumlah_qos,
    persen_qos=persen_qos,
    map_static_file=map_static_file,
    operator_speed_json=json.dumps(operator_speed_json)
    )
    
@app.route("/get_kab/<spt>")
def get_kab(spt):
    selected_year = request.form.get("year", "2026")
    df = load_data(selected_year)
    if spt != "Semua":
        df = df[df["observasi_no_spt"] == spt]
    kab_options = sorted(df["observasi_kota_nama"].dropna().unique().tolist())
    return {"kab_list": kab_options}

@app.route("/get_kec/<spt>/<kab>")
def get_kec(spt, kab):
    selected_year = request.form.get("year", "2026")
    df = load_data(selected_year)
    if spt != "Semua":
        df = df[df["observasi_no_spt"] == spt]
    if kab != "Semua":
        df = df[df["observasi_kota_nama"] == kab]
    kec_options = sorted(df["observasi_kecamatan_nama"].dropna().unique().tolist())
    return {"kec_list": kec_options}

@app.route("/get_cat/<spt>/<kab>/<kec>")
def get_cat(spt, kab, kec):
    selected_year = request.form.get("year", "2026")
    df = load_data(selected_year)
    if spt != "Semua":
        df = df[df["observasi_no_spt"] == spt]
    if kab != "Semua":
        df = df[df["observasi_kota_nama"] == kab]
    if kec != "Semua":
        df = df[df["observasi_kecamatan_nama"] == kec]
    cat_options = sorted(df["scan_catatan"].dropna().unique().tolist())
    return {"cat_list": cat_options}

def rupiah_to_int(x):
    """
    'Rp.1.617.000' -> 1617000
    """
    if pd.isna(x):
        return 0
    return int(re.sub(r"[^\d]", "", str(x))) or 0


def int_to_rupiah(x):
    """
    1617000 -> 'Rp.1.617.000'
    """
    return f"Rp.{x:,.0f}".replace(",", ".")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
