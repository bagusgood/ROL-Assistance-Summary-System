from flask import Flask, render_template_string, request
import requests
import pandas as pd
import plotly.express as px
import plotly
import json
from flask import send_file
import io
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import os
import re


app = Flask(__name__)

# Bersihkan string agar aman untuk nama file
def safe_filename(text):
    return re.sub(r'[\\/*?:"<>|]', "_", str(text).strip())

def load_data():
    url = "https://rol.postel.go.id/api/observasi/allapproved"
    params = {
        "upt": 19,
        "year": "2025",
        "pageIndex": 1,
        "pageSize": 10000
    }
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
        "Accept": "*/*",
        "Referer": "https://rol.postel.go.id/observasi/laporan",
        "X-Requested-With": "XMLHttpRequest",
        "Cookie": "csrf_cookie_name=6701fbfe229f77d47c710eec3391f386; ci_session=1tmigk58v1636foa82v97rmo8o2olsrc"
    }

    try:
        response = requests.get(url, params=params, headers=headers)
        data = response.json().get("data", [])
        return pd.DataFrame(data)
    except Exception as e:
        print("Gagal ambil data API:", e)
        return pd.DataFrame()

@app.route("/download_excel", methods=["POST"])
def download_excel():
    df = load_data()

    # Transformasi jenis identifikasi
    df["observasi_status_identifikasi_name"] = df["observasi_status_identifikasi_name"].str.replace(
        r"OFF AIR \(Sedang Tidak Digunakan\)", "OFF AIR", regex=True
    )
    df['jenis'] = df['observasi_status_identifikasi_name'].apply(
        lambda x: 'Belum Teridentifikasi' if x == 'BELUM DIKETAHUI' else 'Teridentifikasi'
    )

    # Ambil filter dari form
    selected_spt = request.form.get("spt", "Semua")
    selected_kab = request.form.get("kab", "Semua")
    selected_kec = request.form.get("kec", "Semua")
    selected_cat = request.form.get("cat", "Semua")

    # Filter data
    filt = df.copy()
    if selected_spt != "Semua":
        filt = filt[filt["observasi_no_spt"] == selected_spt]
    if selected_kab != "Semua":
        filt = filt[filt["observasi_kota_nama"] == selected_kab]
    if selected_kec != "Semua":
        filt = filt[filt["observasi_kecamatan_nama"] == selected_kec]
    if selected_cat != "Semua":
        filt = filt[filt["scan_catatan"] == selected_cat]

    if filt.empty:
        return "<h3>Data kosong, tidak bisa disimpan.</h3>"
    
    # Ambil nama Kab/Kota yang aktif dari data filter
    kab_aktif = selected_kab if selected_kab != "Semua" else None

    # Ringkasan
    jumlah_kota = filt['observasi_kota_nama'].nunique()
    legal = filt.groupby(['observasi_kota_nama', 'observasi_status_identifikasi_name']).size().reset_index(name='Jumlah')
    band = filt.pivot_table(index=['observasi_kota_nama', 'band_nama', 'jenis'], aggfunc='size', fill_value=0)
    dinas = filt.pivot_table(index=['observasi_kota_nama', 'observasi_service_name', 'jenis'], aggfunc='size', fill_value=0)
    pita = filt.pivot_table(index=['observasi_kota_nama', 'observasi_range_frekuensi', 'observasi_status_identifikasi_name'], aggfunc='size', fill_value=0)

    # Perhitungan tambahan: Identifikasi dan kesesuaian dengan ISR
    try:
        # Load dan sesuaikan nama kolom dari file ISR
        df_ISR = pd.read_csv("Data Target Monitor ISR 2025 - Mataram.csv", on_bad_lines='skip', delimiter=';')
        # Rename kolom dari df_ISR
        df2 = df_ISR.rename(columns={'Freq': 'Frekuensi', 'Clnt Name': 'Identifikasi'})
        
        # Konversi 'Frekuensi' di kedua dataframe ke float64
        filt['observasi_frekuensi'] = pd.to_numeric(filt['observasi_frekuensi'], errors='coerce')
        df2['Frekuensi'] = pd.to_numeric(df2['Frekuensi'], errors='coerce')
        
        # Siapkan freq_df1 dari API
        freq_df1 = filt.groupby(['observasi_frekuensi', 'observasi_sims_client_name']).size().reset_index(name='Jumlah_df1')
        freq_df1 = freq_df1.rename(columns={'observasi_frekuensi': 'Frekuensi', 'observasi_sims_client_name': 'Identifikasi'})
        
        # freq_df2 dari ISR
        freq_df2 = df2.groupby(['Frekuensi', 'Identifikasi']).size().reset_index(name='Jumlah_df2')
        
        # Sekarang merge
        merged = pd.merge(freq_df1, freq_df2, on=['Frekuensi', 'Identifikasi'], how='inner')
        jumlah_sesuai_isr = len(merged)
        print("Jumlah cocok:", jumlah_sesuai_isr)
    except Exception as e:
        print("Gagal menghitung kesesuaian dengan ISR:", e)
        jumlah_sesuai_isr = 0

    # Baca file target_kota.csv dan cocokkan jumlah target ISR untuk kabupaten tersebut
    try:
        # Baca target kota
        df_target_kota = pd.read_csv("target_kota.csv", delimiter=';', on_bad_lines='skip')
        df_target_kota['Kabupaten/Kota'] = df_target_kota['Kabupaten/Kota'].astype(str).str.strip().str.upper()
    
        # Ambil nama kota dari data termonitor
        kota_termonitor = filt['observasi_kota_nama'].dropna().unique()
        kota_termonitor = [str(k).strip().upper() for k in kota_termonitor]
    
        # Cari kecocokan pertama
        jumlah_target_isr = None
        for kota in kota_termonitor:
            match = df_target_kota[df_target_kota['Kabupaten/Kota'] == kota]
            if not match.empty:
                jumlah_target_isr = int(match['Jumlah ISR'].iloc[0])
                break
    
        if jumlah_target_isr:
            persen_sesuai_isr = round((jumlah_sesuai_isr / jumlah_target_isr * 100), 2)
        else:
            persen_sesuai_isr = 0
            print(f"TIDAK DITEMUKAN data target ISR untuk kota: {kota_termonitor}")
    except Exception as e:
        print("Gagal hitung persen target ISR:", e)
        persen_sesuai_isr = 0


    jumlah_iden = len(filt[filt['jenis'] == 'Teridentifikasi'])
    jumlah_total = len(filt)
    persen_teridentifikasi = round((jumlah_iden / jumlah_total * 100), 2) if jumlah_total > 0 else 0

    # Buat Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Ringkasan Laporan"

    def add_centered_row(value, merge_cells=4):
        ws.merge_cells(start_row=ws.max_row + 1, start_column=1, end_row=ws.max_row + 1, end_column=merge_cells)
        cell = ws.cell(row=ws.max_row, column=1, value=value)
        cell.alignment = Alignment(horizontal='center')

    def add_labeled_row(label, value):
        ws.append([label, value])

    # Mulai isi
    add_centered_row("=" * 60)
    add_centered_row("RANGKUMAN HASIL LAPORAN SURAT TUGAS")
    add_centered_row(selected_spt)
    add_centered_row("=" * 60)

    add_labeled_row("Jumlah Kab/Kota Termonitor:", jumlah_kota)
    add_centered_row("=" * 60)

    # Legalitas
    ws.append(["Rangkuman Berdasarkan Legalitas:"])
    for _, row in legal.iterrows():
        ws.append(row.tolist())
    add_centered_row("=" * 60)

    # Band
    ws.append(["Rangkuman Berdasarkan Band:"])
    band_df = band.reset_index()
    for _, row in band_df.iterrows():
        ws.append(row.tolist())
    add_centered_row("=" * 60)

    # Dinas
    ws.append(["Rangkuman Berdasarkan Dinas:"])
    dinas_df = dinas.reset_index()
    for _, row in dinas_df.iterrows():
        ws.append(row.tolist())
    add_centered_row("=" * 60)

    # Pita
    ws.append(["Rangkuman Berdasarkan Pita:"])
    pita_df = pita.reset_index()
    for _, row in pita_df.iterrows():
        ws.append(row.tolist())
    add_centered_row("=" * 60)

    # Tambahan ringkasan
    add_labeled_row("Jumlah Data Rekap:", jumlah_total)
    add_labeled_row("Jumlah Stasiun Radio Teridentifikasi:", f"{jumlah_iden} ({persen_teridentifikasi}%)")
    add_labeled_row("Jumlah Stasiun Radio Sesuai ISR:", f"{jumlah_sesuai_isr} ({persen_sesuai_isr}%)")

    add_centered_row("=" * 60)

    # Atur lebar kolom otomatis
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_len + 2

    # Simpan ke file sementara
    from tempfile import NamedTemporaryFile
    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.seek(0)

    # Ubah nama berdasarkan input
    safe_spt = safe_filename(selected_spt)
    safe_kab = safe_filename(selected_kab)
    safe_kec = safe_filename(selected_kec)
    filename = f"Rekap_Observasi {safe_spt} {safe_kec} {safe_kab}.xlsx"

    return send_file(
        tmp.name,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/", methods=["GET", "POST"])
def index():
    df = load_data()

    # Tambahkan kolom jenis
    df['jenis'] = df['observasi_status_identifikasi_name'].apply(
        lambda x: 'Belum Teridentifikasi' if x == 'BELUM DIKETAHUI' else 'Teridentifikasi'
    )

    if df.empty:
        return "Data tidak tersedia. Periksa koneksi API atau cookie."

    # Semua opsi SPT + "Semua"
    spt_options = df["observasi_no_spt"].dropna().unique().tolist()
    spt_options.insert(0, "Semua")
    selected_spt = request.form.get("spt", "Semua")

    # Filter kab berdasarkan SPT
    if selected_spt == "Semua":
        df_spt = df.copy()
    else:
        df_spt = df[df["observasi_no_spt"] == selected_spt]

    kab_options = df_spt["observasi_kota_nama"].dropna().unique().tolist()
    kab_options.insert(0, "Semua")
    selected_kab = request.form.get("kab", "Semua")

    # Filter kec berdasarkan kab
    if selected_kab == "Semua":
        df_kab = df_spt.copy()
    else:
        df_kab = df_spt[df_spt["observasi_kota_nama"] == selected_kab]

    kec_options = df_kab["observasi_kecamatan_nama"].dropna().unique().tolist()
    kec_options.insert(0, "Semua")
    selected_kec = request.form.get("kec", "Semua")

    # Filter cat berdasarkan kec
    if selected_kec == "Semua":
        df_kec = df_kab.copy()
    else:
        df_kec = df_kab[df_kab["scan_catatan"] == selected_kec]
        
    cat_options = df_kec["scan_catatan"].dropna().unique().tolist()
    cat_options.insert(0, "Semua")
    selected_cat = request.form.get("cat", "Semua")
    
    # Filter akhir untuk grafik
    filt = df.copy()
    if selected_spt != "Semua":
        filt = filt[filt["observasi_no_spt"] == selected_spt]
    if selected_kab != "Semua":
        filt = filt[filt["observasi_kota_nama"] == selected_kab]
    if selected_kec != "Semua":
        filt = filt[filt["observasi_kecamatan_nama"] == selected_kec]

    if filt.empty:
        return f"<h3>Data kosong untuk kombinasi tersebut.</h3><p>SPT: {selected_spt}, Kab: {selected_kab}, Kec: {selected_kec}</p>"

    # Chart
    pie1 = px.pie(filt, names="observasi_status_identifikasi_name", title="Distribusi Legalitas")
    pie_band = px.pie(filt, names="band_nama", title="Distribusi Band")

    bar = filt.groupby(["observasi_service_name", "jenis"]).size().reset_index(name="jumlah").sort_values(by="jumlah", ascending=False)
    total_per_dinas = (
    filt.groupby("observasi_service_name")
    .size()
    .sort_values(ascending=False))
    
    ordered_dinas = total_per_dinas.index.tolist()  # urutan berdasarkan total
    bar1 = px.bar(bar, x="observasi_service_name", y="jumlah", color="jenis", title="Distribusi Dinas & Jenis",
                  labels={"observasi_service_name": "Nama Dinas",
                          "jumlah": "Jumlah Data",
                          "jenis": "Status Identifikasi"},
                  category_orders={"observasi_service_name": ordered_dinas})

    bar_pita = filt.groupby(["observasi_range_frekuensi", "observasi_status_identifikasi_name"]).size().reset_index(name="jumlah").sort_values(by="jumlah", ascending=False)
    bar_pita["pita_singkat"] = bar_pita["observasi_range_frekuensi"].astype(str).str.split('.').str[0]
    total_per_pita = (
    filt.groupby("observasi_range_frekuensi")
    .size()
    .sort_values(ascending=False))
    
    ordered_pita = total_per_pita.index.tolist()  # urutan berdasarkan total
    bar1_pita = px.bar(bar_pita, x="observasi_range_frekuensi", y="jumlah", color="observasi_status_identifikasi_name",
                       title="Distribusi Pita & Legalitas",
                       labels={"observasi_range_frekuensi": "Pita Frekuensi",
                               "jumlah": "Jumlah Data",
                               "observasi_status_identifikasi_name": "Legalitas"},
                       category_orders={"observasi_range_frekuensi": ordered_pita}, hover_name="observasi_range_frekuensi")

    bar1_pita.update_layout(
        xaxis=dict(
            tickmode="array",
            tickvals=bar_pita["observasi_range_frekuensi"],
            ticktext=bar_pita["pita_singkat"]
        )
    )

    pie1_json = json.dumps(pie1, cls=plotly.utils.PlotlyJSONEncoder)
    pie_band_json = json.dumps(pie_band, cls=plotly.utils.PlotlyJSONEncoder)
    bar1_json = json.dumps(bar1, cls=plotly.utils.PlotlyJSONEncoder)
    bar1_pita_json = json.dumps(bar1_pita, cls=plotly.utils.PlotlyJSONEncoder)
    
    return render_template_string('''
    <!DOCTYPE html>
    <html>
    <head>
        <title>DATA OBSERVASI BALAI MONITOR SFR KELAS II MATARAM TAHUN 2025</title>
        <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
        <style>
            body { font-family: Arial; padding: 20px; }
            select { margin-right: 10px; }
            .chart { margin-top: 30px; }
            .chart-row {
                display: flex;
                justify-content: space-between;
                gap: 20px;
                margin-top: 30px;
            }
            .chart {
                flex: 1;
                min-width: 400px;
                height: 500px;
            }
        </style>
    </head>
    <body>
    <div class="header-banner">
        <img src="/static/logo-komdigi.png" alt="Logo Kominfo" class="logo-img">
        <div class="title-text">
            <h1>Rekapitulasi Data Observasi</h1>
            <h3>Balai Monitor SFR Kelas II Mataram - Tahun 2025</h3>
        </div>
    </div>

    <style>
        body {
            font-family: 'Segoe UI', sans-serif;
            background-color: #f4f6f9;
            margin: 0;
            padding: 20px;
        }

        .header-banner {
            display: flex;
            align-items: center;
            gap: 20px;
            padding: 20px;
            background-color: #ffffff;
            border-radius: 10px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        }

        .logo-img {
            height: 60px;
            width: auto;
        }

        .title-text h1 {
            margin: 0;
            font-size: 24px;
            color: #003366;
        }

        .title-text h3 {
            margin: 4px 0 0 0;
            font-size: 16px;
            color: #666;
        }

        .filter-form {
            display: flex;
            flex-wrap: wrap;
            gap: 20px;
            margin: 25px 0;
            background-color: #fff;
            padding: 20px;
            border-radius: 10px;
            box-shadow: 0 2px 6px rgba(0, 0, 0, 0.1);
        }

        .filter-group {
            display: flex;
            flex-direction: column;
            min-width: 200px;
        }

        .filter-group label {
            font-weight: 600;
            margin-bottom: 5px;
            color: #333;
        }

        .filter-group select {
            padding: 8px;
            border: 1px solid #ccc;
            border-radius: 6px;
            font-size: 14px;
        }

        .filter-buttons {
            display: flex;
            flex-direction: column;
            justify-content: flex-end;
            gap: 10px;
        }

        .filter-buttons button {
            padding: 10px 16px;
            background-color: #007bff;
            color: white;
            border: none;
            border-radius: 6px;
            font-weight: bold;
            cursor: pointer;
            transition: 0.3s;
        }

        .filter-buttons button:hover {
            background-color: #0056b3;
        }

        .chart-row {
            display: flex;
            justify-content: space-between;
            gap: 20px;
            margin: 30px 0;
        }

        .chart {
            flex: 1;
            min-width: 400px;
            height: 400px;
        }
    </style>

    <!-- Form utama -->
    <form method="POST" class="filter-form">
        <div class="filter-group">
            <label for="spt">No SPT</label>
            <select name="spt" id="spt">
                {% for spt in spt_options %}
                <option value="{{spt}}" {% if spt == selected_spt %}selected{% endif %}>{{spt}}</option>
                {% endfor %}
            </select>
        </div>

        <div class="filter-group">
            <label for="kab">Kab/Kota</label>
            <select name="kab" id="kab">
                {% for kab in kab_options %}
                <option value="{{kab}}" {% if kab == selected_kab %}selected{% endif %}>{{kab}}</option>
                {% endfor %}
            </select>
        </div>

        <div class="filter-group">
            <label for="kec">Kecamatan</label>
            <select name="kec" id="kec">
                {% for kec in kec_options %}
                <option value="{{kec}}" {% if kec == selected_kec %}selected{% endif %}>{{kec}}</option>
                {% endfor %}
            </select>
        </div>

        <div class="filter-group">
            <label for="cat">Catatan</label>
            <select name="cat" id="cat">
                {% for cat in cat_options %}
                <option value="{{cat}}" {% if cat == selected_cat %}selected{% endif %}>{{cat}}</option>
                {% endfor %}
            </select>
        </div>

        <div class="filter-buttons">
            <button type="submit">🔍 Tampilkan</button>
            <button form="excel-form" type="submit">⬇️ Unduh Rekap</button>
        </div>
    </form>

    <!-- Form Unduh (terpisah) -->
    <form method="POST" action="/download_excel" id="excel-form" style="display:none;">
        <input type="hidden" name="spt" value="{{ selected_spt }}">
        <input type="hidden" name="kab" value="{{ selected_kab }}">
        <input type="hidden" name="kec" value="{{ selected_kec }}">
        <input type="hidden" name="cat" value="{{ selected_cat }}">
    </form>

    <!-- Charts -->
    <div class="chart" id="bar1_pita"></div>

    <div class="chart-row">
        <div class="chart" id="pie1"></div>
        <div class="chart" id="pie_band"></div>
    </div>

    <div class="chart" id="bar1"></div>

    <script>
        Plotly.newPlot("pie1", {{ pie1_json|safe }}.data, {{ pie1_json|safe }}.layout);
        Plotly.newPlot("pie_band", {{ pie_band_json|safe }}.data, {{ pie_band_json|safe }}.layout);
        Plotly.newPlot("bar1", {{ bar1_json|safe }}.data, {{ bar1_json|safe }}.layout);
        Plotly.newPlot("bar1_pita", {{ bar1_pita_json|safe }}.data, {{ bar1_pita_json|safe }}.layout);
    </script>
</body>

    </html>
    ''',
    spt_options=spt_options, kab_options=kab_options, kec_options=kec_options, cat_options=cat_options,
    selected_spt=selected_spt, selected_kab=selected_kab, selected_kec=selected_kec, selected_cat=selected_cat,
    pie1_json=pie1_json, pie_band_json=pie_band_json, bar1_json=bar1_json, bar1_pita_json=bar1_pita_json)

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=1346)


